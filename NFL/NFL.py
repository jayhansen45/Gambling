"""
Next Steps:

Add some shit about the next game?
    Email self when the next game is so can run before
    Set reminder would be even better

Comments

"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date
import time
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
#chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)


filedate=date.today()
day = (filedate.strftime('%d/%m/%Y'))
day_site = (filedate.strftime('%A'))
filedate = (filedate.strftime('%A %#d %B %Y'))
tomorrow = date.today()+timedelta(days = 1)
tomorrow_site = (tomorrow.strftime('%a'))


#Finds the file stored in the location and saves the working sheet and saves a copy

#Personal Laptop
"""
filename = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\NFL\\NFL Tracker.xlsx"
newLocation = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\NFL\\Historical\\NFL Tracker.xlsx"
newName = "NFL " + filedate + ".xlsx"
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\Historical\\NFL Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\Historical\\" + newName)
"""

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\NFL Tracker.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\Historical\\NFL Tracker.xlsx"
newName = "NFL " + filedate + ".xlsx"
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\Historical\\NFL Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NFL\\Historical\\" + newName)


workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
odds_row=1
    
for odds_row in range(1, 1048576):
    if sheet.cell(odds_row, 9).value is None:
        break
odds_row=odds_row-1

score_row=1
    
for score_row in range(1, 1048576):
    if sheet.cell(score_row, 4).value is None:
        break
score_row=score_row-1

#Finds current round number
prev_round = 0

for previous_round in range(1, 1048576):
    if sheet.cell(previous_round, 1).value is None:
        break

if prev_round != 1:
    round_num = sheet.cell(previous_round-1, 1).value
else:
    round_num = 1

if day_site == "Thursday":
    round_num = round_num + 1



#Find out how many games are being played the next day
site = 'https://www.footballdb.com/scores/index.html?lg=NFL&yr=2022&type=reg&wk=' + str(round_num)
driver.get(site)

time.sleep(5)
games_num = 0

games = driver.find_elements(By.XPATH, '//*[@class="left"]')

for a in range(0, len(games)):
    if games[a].text.split(',')[0] == day_site:
        games_num = games_num + 1


#Finds scores from the previous day
if day_site == "Tuesday":
    webpage_response = requests.get('https://sports.yahoo.com/nfl/scoreboard/?confId=&dateRange=' + str(round_num) + '&schedState=2&scoreboardSeason=2022')
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")

    messy_teams_scores = soup.find_all(attrs={'class':'Fw(n) Fz(12px)'})
    messy_scores = soup.find_all(attrs={'class':'YahooSans Fw(700)! Va(m) Fz(24px)!'})

    teams_scores = []
    scores = []
    difference = odds_row-score_row


    for a in range(0, len(messy_teams_scores)):
        teams_scores.append(messy_teams_scores[a].get_text())

    print(teams_scores)
    

    for a in range(0, len(messy_scores)):
        scores.append(messy_scores[a].get_text())

    print(scores)

    for i in range(score_row+1, odds_row+1):
        temp = sheet.cell(i, 2).value.split()
        for j in range(0, len(teams_scores)-1):
            if temp[len(temp)-1] == teams_scores[j]:
                sheet.cell(i, 4).value = int(scores[j-2*games_num])
                sheet.cell(i, 5).value = int(scores[j+1-2*games_num])



print("Games Tomorrow = ", games_num)
if games_num > 0:

    #Finds odds for outright wins
    #-------------------------------
    #Website that has the teams and odds
    webpage_response = requests.get('https://www.sportsbet.com.au/betting/american-football/nfl')
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")
    
    
    #Pulls out the teams and odds
    messy_teams_JUST = soup.find_all(attrs={'class':'size14_f7opyze Endeavour_fhudrb0 medium_f1wf24vo participantText_fivg86r'})
    messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})
    

    #Finds big win and little win odds
    #----------------------------------
    #Stores website in the web driver
    driver.get('https://www.sportsbet.com.au/betting/american-football/nfl')

    bets = ["Total Match Points", "Match Betting", "Big Win Little Win", "Double Result", "Team to Score Last", "Tri-Bet", "Tri-Bet 2", "Race To 10", "Race To 15", "Race To 20", "Race To 25", "Race To 30", "Both Teams to Score 10 Points", "Both Teams to Score 15 Points", "Both Teams to Score 20 Points", "Both Teams to Score 25 Points", "Both Teams to Score 30 Points", "Both Teams to Score 35 Points", "Both Teams to Score 40 Points"]

    messy_more_odds = []
    odds =[]
    count = 5
    row = 2
    teams_JUST = []

    for a in messy_teams_JUST:
        teams_JUST.append(a.get_text())


    #Finds the first row that hasn't been used
    m=1
        
    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break
    m=m-2

    for i in range(0, games_num):
        sheet.cell(i+2+m, 2).value = teams_JUST[2*i]
        sheet.cell(i+2+m, 3).value = teams_JUST[2*i+1]
        sheet.cell(i+2+m, 1).value = round_num

    more_odds = []

    for i in range(0, len(bets)):
        print(bets[i])
        row = 2
        teams =[]

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[i] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, 'competition-event-participant')]")
        for a in messy_teams:
            teams.append(a.text)
        
        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []
        odds = []
        messy_total_points = []
        total_points = []
        

        for a in messy_more_odds:
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))
        total_games = len(teams)/3


        if bets[i] == "Total Match Points":
            messy_total_points = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, '-outcome-outcome-name')]")
            j = 0
            for a in range(0, len(messy_total_points), 2):
                    temp = messy_total_points[a].text
                    temp2 = temp.split("(")
                    temp3 = temp2[1].split(")")
                    total_points.append(float(temp3[0]))
            
            for row in range(2, games_num + 2):
                sheet.cell(row+m, 6).value = total_points[row-2]
                
        if len(more_odds) == total_games*2:
            j = 0
            for row in range(2, games_num + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                j=j+2
            count = count + 2

        elif len(more_odds) == total_games*3:
            j = 0
            for row in range(2, games_num + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                j=j+3
            count = count+3
            
        elif len(more_odds) == total_games*4:
            j = 0
            for row in range(2, games_num + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                j=j+4
            count = count+4
            
        elif len(more_odds) == total_games*5:
            j = 0
            for row in range(2, games_num + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                sheet.cell(row+m, count+4+2).value = more_odds[j+4]
                j=j+5
            count = count+5
        print("Done")

workbook.save("NFL Tracker.xlsx")
driver.close()
