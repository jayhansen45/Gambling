""" TO DO

Comments

Fix up what the numbers are for each stat

Create new file that pulls out the odds
"""


import openpyxl as xl
import requests
import re
import string
import statistics as st
from scipy.stats import norm
from bs4 import BeautifulSoup

site = "https://www.pro-football-reference.com/players/"
pass_yds_stat = []
rush_yds_stat = []
pass_td_stat = []
rush_td_stat = []
site_array = []
row = 2

workbook = xl.load_workbook("2022 Season Data.xlsx")
pass_yds_sheet = workbook.worksheets[0]
rush_yds_sheet = workbook.worksheets[1]
pass_td_sheet = workbook.worksheets[2]
rush_td_sheet = workbook.worksheets[3]

for i in range(0, 26):
    if i !=23:
        site_array.append(site+chr(i+65)+"/")

for j in range(0, len(site_array)):
    print(site_array[j])
    webpage_response = requests.get(site_array[j])
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")
    bold_links = soup.find_all('b')
    print(len(bold_links))
    for k in range(0, len(bold_links)):
        if bold_links[k].find('a') != None:
            temp = bold_links[k].find('a').get('href').split('.html')
            site = "https://www.pro-football-reference.com/" + temp[0] + "/gamelog/2021"
            webpage_response2 = requests.get(site)
            webpage2 = webpage_response2.content
            soup2 = BeautifulSoup(webpage2, "html.parser")
            pass_yds = soup2.find_all(attrs={'data-stat':'pass_yds'})
            rush_yds = soup2.find_all(attrs={'data-stat':'rush_yds'})
            pass_td = soup2.find_all(attrs={'data-stat':'pass_td'})
            rush_td = soup2.find_all(attrs={'data-stat':'rush_td'})

            name = bold_links[k].find('a').text
            print(name)

            for h in range(0, len(pass_yds)):
                if pass_yds[h].text != "Yds":
                    pass_yds_stat.append(int(pass_yds[h].text))
            print(pass_yds_stat)
            
            for h in range(0, len(rush_yds)):
                if rush_yds[h].text != "Yds":
                    if rush_yds[h].text == '':
                        rush_yds_stat.append(0)
                    else:
                        rush_yds_stat.append(int(rush_yds[h].text))
            print(rush_yds_stat)   

            for h in range(0, len(pass_td)):
                if pass_td[h].text != "TD":
                    pass_td_stat.append(int(pass_td[h].text))
            print(pass_td_stat)

            for h in range(0, len(rush_td)):
                if rush_td[h].text != "TD":
                    if rush_td[h].text == '':
                        rush_td_stat.append(0)
                    else:
                        rush_td_stat.append(int(rush_td[h].text))                   
            print(rush_td_stat)

            if len(pass_yds)>1:
                print(name)
                pass_yds_average = (sum(pass_yds_stat)/len(pass_yds_stat))
                pass_yds_std = (st.stdev(pass_yds_stat))
                
                pass_yds_15_percentage = (1-norm.cdf(14, pass_yds_average, pass_yds_std))*100
                pass_yds_15_req_odds = 1/(pass_yds_15_percentage/100)

                pass_yds_20_percentage = (1-norm.cdf(19, pass_yds_average, pass_yds_std))*100
                pass_yds_20_req_odds = 1/(pass_yds_20_percentage/100)

                pass_yds_25_percentage = (1-norm.cdf(24, pass_yds_average, pass_yds_std))*100
                pass_yds_25_req_odds = 1/(pass_yds_25_percentage/100)
                
                pass_yds_sheet.cell(row, 1).value = name
                
                pass_yds_sheet.cell(row, 2).value = pass_yds_average
                pass_yds_sheet.cell(row, 3).value = pass_yds_std
                
                pass_yds_sheet.cell(row, 4).value = pass_yds_15_percentage
                pass_yds_sheet.cell(row, 5).value = pass_yds_15_req_odds
                
                pass_yds_sheet.cell(row, 6).value = pass_yds_20_percentage
                pass_yds_sheet.cell(row, 7).value = pass_yds_20_req_odds
                
                pass_yds_sheet.cell(row, 8).value = pass_yds_25_percentage
                pass_yds_sheet.cell(row, 9).value = pass_yds_25_req_odds
                
            if len(rush_yds_stat)>1:
                rush_yds_average = (sum(rush_yds_stat)/len(rush_yds_stat))
                rush_yds_std = (st.stdev(rush_yds_stat))
                
                rush_yds_15_percentage = (1-norm.cdf(4, rush_yds_average, rush_yds_std))*100
                rush_yds_15_req_odds = 1/(rush_yds_15_percentage/100)

                rush_yds_20_percentage = (1-norm.cdf(6, rush_yds_average, rush_yds_std))*100
                rush_yds_20_req_odds = 1/(rush_yds_20_percentage/100)
                
                rush_yds_25_percentage = (1-norm.cdf(8, rush_yds_average, rush_yds_std))*100
                rush_yds_25_req_odds = 1/(rush_yds_25_percentage/100)
                
                rush_yds_sheet.cell(row, 1).value = name
                
                rush_yds_sheet.cell(row, 2).value = rush_yds_average
                rush_yds_sheet.cell(row, 3).value = rush_yds_std
                
                rush_yds_sheet.cell(row, 4).value = rush_yds_15_percentage
                rush_yds_sheet.cell(row, 5).value = rush_yds_15_req_odds
                
                rush_yds_sheet.cell(row, 6).value = rush_yds_20_percentage
                rush_yds_sheet.cell(row, 7).value = rush_yds_20_req_odds
                
                rush_yds_sheet.cell(row, 8).value = rush_yds_25_percentage
                rush_yds_sheet.cell(row, 9).value = rush_yds_25_req_odds

                

            if len(pass_td_stat)>1:
                pass_td_average = (sum(pass_td_stat)/len(pass_td_stat))
                pass_td_std = (st.stdev(pass_td_stat))
                
                pass_td_15_percentage = (1-norm.cdf(4, pass_td_average, pass_td_std))*100
                pass_td_15_req_odds = 1/(pass_td_15_percentage/100)

                pass_td_20_percentage = (1-norm.cdf(6, pass_td_average, pass_td_std))*100
                pass_td_20_req_odds = 1/(pass_td_20_percentage/100)

                pass_td_25_percentage = (1-norm.cdf(8, pass_td_average, pass_td_std))*100
                pass_td_25_req_odds = 1/(pass_td_25_percentage/100)
                
                pass_td_sheet.cell(row, 1).value = name
                
                pass_td_sheet.cell(row, 2).value = pass_td_average
                pass_td_sheet.cell(row, 3).value = pass_td_std
                
                pass_td_sheet.cell(row, 4).value = pass_td_15_percentage
                pass_td_sheet.cell(row, 5).value = pass_td_15_req_odds
                
                pass_td_sheet.cell(row, 6).value = pass_td_20_percentage
                pass_td_sheet.cell(row, 7).value = pass_td_20_req_odds
                
                pass_td_sheet.cell(row, 8).value = pass_td_25_percentage
                pass_yd_sheet.cell(row, 9).value = pass_yd_25_req_odds

                row = row+1

            if len(rush_td_stat)>1:
                rush_td_average = (sum(rush_td_stat)/len(rush_td_stat))
                rush_td_std = (st.stdev(rush_td_stat))
                
                rush_td_15_percentage = (1-norm.cdf(4, rush_td_average, rush_td_std))*100
                rush_td_15_req_odds = 1/(rush_td_15_percentage/100)

                rush_td_20_percentage = (1-norm.cdf(6, rush_td_average, rush_td_std))*100
                rush_td_20_req_odds = 1/(rush_td_20_percentage/100)

                rush_td_25_percentage = (1-norm.cdf(8, rush_td_average, rush_td_std))*100
                rush_td_25_req_odds = 1/(rush_td_25_percentage/100)
                
                rush_td_sheet.cell(row, 1).value = name
                
                rush_td_sheet.cell(row, 2).value = rush_td_average
                rush_td_sheet.cell(row, 3).value = rush_td_std
                
                rush_td_sheet.cell(row, 4).value = rush_td_15_percentage
                rush_td_sheet.cell(row, 5).value = rush_td_15_req_odds
                
                rush_td_sheet.cell(row, 6).value = rush_td_20_percentage
                rush_td_sheet.cell(row, 7).value = rush_td_20_req_odds
                
                rush_td_sheet.cell(row, 8).value = rush_td_25_percentage
                rush_td_sheet.cell(row, 9).value = rush_td_25_req_odds

                row = row+1


            pass_yds_stat = []
            rush_yds_stat = []
            pass_td_stat = []
            rush_td_stat = []

    workbook.save("2022 Season Data.xlsx")

