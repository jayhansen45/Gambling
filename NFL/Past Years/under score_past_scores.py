import openpyxl as xl
import requests
import re
import string
import statistics as st
from scipy.stats import norm
from bs4 import BeautifulSoup

year  = []
full_scores = []
book = xl.Workbook()
sheet = book.worksheets[0]
col = 2
row = 2

for i in range (2010, 2022):
    year = []
    for j in range (1, 18):
        site = "https://www.pro-football-reference.com/years/" + str(i) + "/week_" + str(j) + ".htm"
        webpage_response = requests.get(site)
        webpage = webpage_response.content
        soup = BeautifulSoup(webpage, "html.parser")
        wins = 0
        year = []
        sheet.cell(1, i-2008).value = i
        game_summary = soup.find_all(attrs={'class':'game_summary expanded nohover'})
        for a in game_summary:
            temp = a.find_all(attrs={'class':'right'})
            for b in range(0, 3):
                if (temp[b].text).isnumeric() == True:
                    year.append(temp[b].text)
        print(year)
        for p in range(0, 2*len(game_summary), 2):
            sheet.cell(j+1, 1).value = j
            difference = abs(int(year[p])-int(year[p+1]))
            if difference < 8:
                wins = wins + 2.11
        sheet.cell(j+1, i-2008).value = wins
        row = row + len(game_summary)

book.save("Past Years.xlsx")


            

