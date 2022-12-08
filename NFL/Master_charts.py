"""


"""



import openpyxl as xl
import os
from openpyxl.chart import (LineChart, Reference, Series)
from string import ascii_uppercase

#Opens and creates the relevant spreadsheets
data ="C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\EPL Tracker.xlsx"
#data ="C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Baseball_Data.xlsx"
data_book = xl.load_workbook(data, data_only=True)
new_book = xl.Workbook()
new_book_data = new_book.worksheets[0]
new_book_data.title = "Data"
new_book_charts = new_book.create_sheet()
new_book_charts.title = "Charts"
analysis_sheet = data_book.worksheets[2]
summary_sheet = data_book.worksheets[0]
"""

epl_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\EPL Tracker.xlsx"
epl_book = xl.load_workbook(epl_file, data_only = True)
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\EPL Tracker.xlsx"')

bundesliga_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Bundesliga Tracker.xlsx"
bundesliga_book = xl.load_workbook(bundesliga_file, data_only = True)
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Bundesliga Tracker.xlsx')

la_liga_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\La Liga Tracker.xlsx"
la_liga_book = xl.load_workbook(la_liga_file, data_only = True)
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\La Liga Tracker.xlsx')

ligue_1_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Ligue 1 Tracker.xlsx"
ligue_1_book = xl.load_workbook(ligue_1_file, data_only = True)
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Ligue 1 Tracker.xlsx"')

serie_a_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Serie A Tracker.xlsx"
serie_a_book = xl.load_workbook(serie_a_file, data_only = True)
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Serie A Tracker.xlsx')

"""
master_file = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Master.xlsx"
master_book = xl.load_workbook(master_file, data_only = True)
#os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Master.xlsx"')

master_sheet = master_book.worksheets[0]

for i in range(4, 24):
    k = 1
    count = 0
    multiplier = 0
    running_multiplier = 0
    total_count = 0
    wins = 0
    total_wins = 0
    file = master_sheet.cell(i, 2).value
    book = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\" + file + " Tracker.xlsx", data_only = True)
    sheet = book.worksheets[2]
    rounds_sheet = book.worksheets[0]
    bet = master_sheet.cell(i, 3).value
    j = 1
    
    while sheet.cell(1, j).value != bet:
            j = j + 1

    while (rounds_sheet.cell(k, 17).value is None) == False:
        k = k + 1

    rounds = k - 1
    row = 7
    round_num = rounds_sheet.cell(rounds, 17).value

    for h in range(0, round_num):
        this_round = sheet.cell(row, 1).value
        while sheet.cell(row, 1).value == this_round:
            count = count + sheet.cell(row, j-1).value
            total_count = total_count + sheet.cell(row, j-1).value
            wins = wins + sheet.cell(row, j).value
            total_wins = total_wins + sheet.cell(row, j).value
            row = row + 1
        if count > 0:
            multiplier = wins/count
            running_multiplier = total_wins/total_count
        else:
            multiplier = 0
        wins = 0
        count = 0
        print(file, bet, round_num, multiplier, running_multiplier)
        new_book_data.cell(h+2, 1).value = h+1
        new_book_data.cell(h+2, 2).value = 1
        new_book_data.cell(1, 2*i-5).value = file + " " + bet
        new_book_data.cell(h+2, 2*i-5).value = multiplier
        new_book_data.cell(h+2, 2*i-4).value = running_multiplier

k = 9

#Creates the charts
for a in range(1, 21):

    chart = LineChart()
    chart.title = master_sheet.cell(a+3, 2).value + " " + master_sheet.cell(a+3, 3).value
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data = Reference(new_book_data, min_col = 2*a+1, min_row = 2, max_col = 2*a+1, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "B"+str(a*18-16))

    chart = LineChart()
    chart.title = master_sheet.cell(a+3, 2).value + " " + master_sheet.cell(a+3, 3).value
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data2 = Reference(new_book_data, min_col = 2*a+2, min_row = 2, max_col = 2*a+2, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data2)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "K"+str(a*18-16))

new_book.active = new_book["Charts"]
new_book.save("Charts.xlsx")
os.system('start "excel" "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Charts.xlsx"')

