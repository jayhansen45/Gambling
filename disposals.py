""" TO DO
Maybe outlier if I can be fucked

Fix for when it isn't 13 games

Try again with autosizing

Add comments below

"""

import openpyxl as xl
import string
import requests
import statistics
import numpy as np
import warnings
from scipy.stats import norm
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore")
urls = ['https://afltables.com/afl/stats/teams/adelaide/2022_gbg.html', 'https://afltables.com/afl/stats/teams/brisbanel/2022_gbg.html', 'https://afltables.com/afl/stats/teams/carlton/2022_gbg.html', 'https://afltables.com/afl/stats/teams/collingwood/2022_gbg.html', 'https://afltables.com/afl/stats/teams/essendon/2022_gbg.html', 'https://afltables.com/afl/stats/teams/fremantle/2022_gbg.html', 'https://afltables.com/afl/stats/teams/geelong/2022_gbg.html', 'https://afltables.com/afl/stats/teams/goldcoast/2022_gbg.html', 'https://afltables.com/afl/stats/teams/gws/2022_gbg.html', 'https://afltables.com/afl/stats/teams/hawthorn/2022_gbg.html', 'https://afltables.com/afl/stats/teams/melbourne/2022_gbg.html', 'https://afltables.com/afl/stats/teams/kangaroos/2022_gbg.html', 'https://afltables.com/afl/stats/teams/padelaide/2022_gbg.html', 'https://afltables.com/afl/stats/teams/richmond/2022_gbg.html', 'https://afltables.com/afl/stats/teams/stkilda/2022_gbg.html', 'https://afltables.com/afl/stats/teams/bullldogs/2022_gbg.html', 'https://afltables.com/afl/stats/teams/westcoast/2022_gbg.html', 'https://afltables.com/afl/stats/teams/swans/2022_gbg.html']
depth = 0
games = 13

workbook = xl.Workbook()
worksheet = workbook.worksheets[0]

worksheet.cell(1, 1).value = "Team"
worksheet.column_dimensions['B'].width = 15
worksheet.cell(1, 2).value = "Players"
worksheet.column_dimensions['A'].width = 15
for i in range(3, games+3):
    worksheet.cell(1, i).value = "Round " + str(i-2)
worksheet.cell(1, i+1).value = "Average"
worksheet.cell(1, i+2).value = "StDev"
worksheet.cell(1, i+3).value = "15+ Percentage"
worksheet.cell(1, i+4).value = "15+ Req Odds"
worksheet.cell(1, i+5).value = "15+ Odds"
worksheet.cell(1, i+6).value = "15+ Difference"

worksheet.cell(1, i+7).value = "20+ Percentage"
worksheet.cell(1, i+8).value = "20+ Req Odds"
worksheet.cell(1, i+9).value = "20+ Odds"
worksheet.cell(1, i+10).value = "20+ Difference"

worksheet.cell(1, i+11).value = "25+ Percentage"
worksheet.cell(1, i+12).value = "25+ Req Odds"
worksheet.cell(1, i+13).value = "25+ Odds"
worksheet.cell(1, i+14).value = "25+ Difference"

worksheet.cell(1, i+15).value = "30+ Percentage"
worksheet.cell(1, i+16).value = "30+ Req Odds"
worksheet.cell(1, i+17).value = "30+ Odds"
worksheet.cell(1, i+18).value = "30+ Difference"

worksheet.cell(1, i+19).value = "35+ Percentage"
worksheet.cell(1, i+20).value = "35+ Req Odds"
worksheet.cell(1, i+21).value = "35+ Odds"
worksheet.cell(1, i+22).value = "35+ Difference"
worksheet.cell(1, i+23).value = "Max Difference"

j=i
i=0

for i in range(1, j+23+1):
    if i > 26:
        worksheet.column_dimensions['A'+ string.ascii_uppercase[i-1-26]].width = 15       
    else:
        worksheet.column_dimensions[string.ascii_uppercase[i-1]].width = 15


for site in urls:
    print(site)
    webpage_response = requests.get(site)
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")

    table = soup.find_all(attrs={'class':'sortable'})
    team = soup.find('h2').text
    team = team.split('-')

    
    messy_disposals = table[0].find_all('td')
    messy_tog = table[21].find_all('td')

    disposals = []
    tog = []

    for a in messy_disposals:
        if (a.text == '\xa0' or a.text == '-'):
            a = '0'
            disposals.append(a)
        else:
            disposals.append(a.text)

    for a in messy_tog:
        if (a.text == '\xa0' or a.text == '-'):
            a = '0'
            tog.append(a)
        else:
            tog.append(a.text)

    i = 1

    while disposals[i].isnumeric():
        i=i+1

    games = i-2
    
    gross_players = []

    i=0

    while (i < len(disposals)):
        gross_players.append(disposals.pop(i))
        tog.pop(i)
        i=i+games+1

    players = []

    for i in range(0, len(gross_players)):
        gross_players[i] = gross_players[i].split(', ')
        players.append(gross_players[i][1] + " " + gross_players[i][0])

    i=games

    while (i < len(disposals)):
        disposals.pop(i)
        tog.pop(i)
        i=i+games

    b=0
    c=0



    disposals_numpy = np.array(disposals)

    disp = disposals_numpy.reshape(len(players), games)

    tog_numpy = np.array(tog)

    time = tog_numpy.reshape(len(players), games)    

    sum = 0


    for a in range(0, len(players)):
        sum = 0
        count = 0
        values = []
        worksheet.cell(a+2+depth, 1).value = team[0]
        worksheet.cell(a+2+depth, 2).value = players[a]
        for t in range(0, games):
            worksheet.cell(a+2+depth, t+3).value = int(disp[a][t])
            if int(time[a][t]) < 60:
                worksheet.cell(a+2+depth, t+3).value = 0
                
            if worksheet.cell(a+2+depth, t+3).value != 0:
                sum = sum + int(disp[a][t])
                count = count + 1
                values.append(int(disp[a][t]))
        if count > 6:
            avg = sum/count
        else:
            avg = 0

        if games < 13:
            t = t + 1
        worksheet.cell(a+2+depth, t+4).value = avg
        values_numpy = np.array(values)
        worksheet.cell(a+2+depth, t+5).value = np.std(values_numpy)

        worksheet.cell(a+2+depth, t+6).value = (1-norm.cdf(14, avg, np.std(values_numpy)))*100
        worksheet.cell(a+2+depth, t+7).value = 1/(worksheet.cell(a+2+depth, t+6).value/100)
        worksheet.cell(a+2+depth, t+9).value = "=T"+str((a+2+depth))+"-"+"S"+str(a+2+depth)

        
        worksheet.cell(a+2+depth, t+10).value = (1-norm.cdf(19, avg, np.std(values_numpy)))*100
        worksheet.cell(a+2+depth, t+11).value = 1/(worksheet.cell(a+2+depth, t+10).value/100)
        worksheet.cell(a+2+depth, t+13).value = "=X"+str((a+2+depth))+"-"+"W"+str(a+2+depth)

        worksheet.cell(a+2+depth, t+14).value = (1-norm.cdf(24, avg, np.std(values_numpy)))*100
        worksheet.cell(a+2+depth, t+15).value = 1/(worksheet.cell(a+2+depth, t+14).value/100)
        worksheet.cell(a+2+depth, t+17).value = "=AB"+str((a+2+depth))+"-"+"AA"+str(a+2+depth)

        worksheet.cell(a+2+depth, t+18).value = (1-norm.cdf(29, avg, np.std(values_numpy)))*100
        worksheet.cell(a+2+depth, t+19).value = 1/(worksheet.cell(a+2+depth, t+18).value/100)
        worksheet.cell(a+2+depth, t+21).value = "=AF"+str((a+2+depth))+"-"+"AE"+str(a+2+depth)
        
        worksheet.cell(a+2+depth, t+22).value = (1-norm.cdf(34, avg, np.std(values_numpy)))*100
        worksheet.cell(a+2+depth, t+23).value = 1/(worksheet.cell(a+2+depth, t+22).value/100)
        worksheet.cell(a+2+depth, t+25).value = "=AJ"+str((a+2+depth))+"-"+"AI"+str(a+2+depth)

        if avg > 0:
            worksheet.cell(a+2+depth, t+26).value = ("=MAX(U" + str(a+2+depth) + ", Y" + str(a+2+depth) + ", AC" + str(a+2+depth) + ", AG" + str(a+2+depth) + ", AK" + str(a+2+depth) + ")")
        else:
            worksheet.cell(a+2+depth, t+26).value = 0

    depth = depth + len(players)


        
workbook.save("Disposals Tracking.xlsx")
