"""
Next Steps:
Fix scores for if there is a double header
    Do a check to see if there is already that same team and if so go for the second one
Delete the pitcher from the team names
Fix for when a game doesn't have odds yet
    Also pull the scores for each type of bet and match the the odds according to this
Sort out the tracking of other stats from games

"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date
import time
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='111.0.5563.64').install(), options = chrome_options)

filedate=date.today()
day = (filedate.strftime('%d/%m/%Y'))
day_site = (filedate.strftime('%A'))
filedate = (filedate.strftime('%A %#d %B %Y'))
tomorrow = date.today()+timedelta(days = 1)
tomorrow_site = (tomorrow.strftime('%a'))

#Personal Laptop
"""
filename = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\MLB\\MLB Tracker.xlsx"
newLocation = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\MLB\\Historical\\MLB Tracker.xlsx"
newName = "MLB " + filedate + ".xlsx"
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\MLB Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\" + newName)
"""

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\MLB Tracker.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\MLB Tracker.xlsx"
newName = "MLB " + filedate
shutil.copyfile(filename, newLocation)
os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\MLB Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\" + newName + ".xlsx")

workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 2).value is None:
        break
m=m-1

#Gets the scores for the previous day
#------------------------------------

#Finds the first cell without a score
t=1
    
for t in range(1, 1048576):
    if sheet.cell(t, 4).value is None:
        break
t=t

#Gets yesterdays date
today=date.today()
yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')

#Website to get the scores from
webpage_response = requests.get('https://sports.yahoo.com/mlb/scoreboard/?confId=&schedState=2&dateRange=', yesterday)
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the text for the scores, mascots and teams
messy_scores = soup.find_all(attrs={'class':'YahooSans Fw(700)! Va(m) Fz(24px)!'})
messy_odds_mascots = soup.find_all(attrs={'class':'Fw(n) Fz(12px)'})
messy_odds_teams = soup.find_all(attrs={'class':'YahooSans Fw(700)! Fz(14px)!'})
scores = []
odds_teams = []
mascots = []

#Pulls out just the text and stores in list
for a in messy_scores:
    scores.append(a.text)

for a in messy_odds_teams:
    odds_teams.append(a.text)

for a in messy_odds_mascots:
    mascots.append(a.text)                                 

for a in range(0, len(mascots)):
    odds_teams[a] = odds_teams[a] + " " + mascots[a]

odds_both = []

#Creates single list that has scores and teams
for i in range(0, len(scores)):
    odds_both.append([])
    odds_both[i].append(odds_teams[i])
    odds_both[i].append(scores[i])


#Saves the scores in the correct cell based on the associated team
for a in range(0, m):
    for k in range(0, len(odds_both)):
        if ((sheet.cell(a+t, 2).value == odds_both[k][0]) and (sheet.cell(a+t, 4).value is None)):
            for h in range(0, 2):
                sheet.cell(a+t, 4+h).value = int(odds_both[k+h][1])
                sheet.cell(a+t, 1).value = today
                sheet.cell(a+t, 1).number_format = 'dd/mm/yyyy'


#Goes through and deletes rows that don't have a score. Postponed games etc
for a in range(1, m):
    if (sheet.cell(a, 4).value is None):
        sheet.delete_rows(a, 1)

workbook.save("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\MLB\\Historical\\" + newName + " - Results.xlsx")

games_num = 14

if games_num > 0:

    #Finds odds for outright wins
    #-------------------------------
    #Website that has the teams and odds
    webpage_response = requests.get('https://www.sportsbet.com.au/betting/baseball/mlb-matches')
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")
    
    
    #Pulls out the teams and odds
    messy_teams_JUST = soup.find_all(attrs={'class':'size14_f7opyze Endeavour_fhudrb0 medium_f1wf24vo participantText_fivg86r'})
    messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})
    

    #Finds big win and little win odds
    #----------------------------------
    #Stores website in the web driver
    driver.get('https://www.sportsbet.com.au/betting/baseball/mlb-matches')

    bets = ["Total Runs", "Money Line", "Big Win Little Win", "Double Result", "Lead after 3rd Inning", "Lead after 6th Inning", "Most Hits", "Race to 3 Runs", "Race to 4 Runs", "Race to 5 Runs", "Race to 6 Runs", "Team with Highest Scoring Inning", "Team to Score Last Wins Game", "Team to Score Last", "Tri-Bet", "Tri-Bet 2", "Tri-Bet 3"]

    messy_more_odds = []
    odds =[]
    count = 5
    row = 2
    teams_JUST = []

    for a in messy_teams_JUST:
        teams_JUST.append(a.get_text())


    #Finds the first row that hasn't been used
    m=1
        
    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break
    m=m-2

    for i in range(0, games_num):
        sheet.cell(i+2+m, 2).value = teams_JUST[2*i]
        sheet.cell(i+2+m, 3).value = teams_JUST[2*i+1]

    more_odds = []

    for i in range(0, len(bets)):
        print(bets[i])
        row = 2
        teams =[]

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[i] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, 'competition-event-participant')]")
        for a in messy_teams:
            teams.append(a.text)
        
        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []
        odds = []
        messy_total_points = []
        total_points = []
        

        for a in messy_more_odds:
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))

        total_games = len(teams)/3


        if bets[i] == "Total Runs":
            messy_total_points = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, '-outcome-outcome-name')]")
            j = 0
            for a in range(0, len(messy_total_points), 2):
                    temp = messy_total_points[a].text
                    temp2 = temp.split("(")
                    temp3 = temp2[1].split(")")
                    total_points.append(float(temp3[0]))
            
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, 6).value = total_points[row-2]
                
        if len(more_odds) == total_games*2:
            j = 0
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                j=j+2
            count = count + 2

        elif len(more_odds) == total_games*3:
            j = 0
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                j=j+3
            count = count+3
            
        elif len(more_odds) == total_games*4:
            j = 0
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                j=j+4
            count = count+4
            
        elif len(more_odds) == total_games*5:
            j = 0
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                sheet.cell(row+m, count+4+2).value = more_odds[j+4]
                j=j+5
            count = count+5
            
        elif len(more_odds) == total_games*6:
            j = 0
            for row in range(2, int(total_games) + 2):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                sheet.cell(row+m, count+4+2).value = more_odds[j+4]
                sheet.cell(row+m, count+5+2).value = more_odds[j+5]
                j=j+6
            count = count+6
        print("Done")


workbook.save("MLB Tracker.xlsx")
driver.quit()

