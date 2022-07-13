"""
Next Steps:
Fix scores for if there is a double header
    Do a check to see if there is already that same team and if so go for the second one


"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome Beta\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(service=Service(ChromeDriverManager(version='104.0.5112.20').install()), options = chrome_options)

filedate=datetime.today()
filedate = (filedate.strftime('%Y-%m-%d'))

#Finds the file stored in the location and saves the working sheet and saves a copy
filename ="C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Baseball_Data.xlsx"
newLocation = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\Baseball_Data.xlsx"
newName = "Baseball_Data_" + filedate + ".xlsx"
shutil.copyfile(filename, newLocation)
os.rename("C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\Baseball_Data.xlsx", "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\" + newName)
workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break
m=m-1

#Gets the scores for the previous day
#------------------------------------

#Finds the first cell without a score
t=1
    
for t in range(1, 1048576):
    if sheet.cell(t, 9).value is None:
        break
t=t

#Gets yesterdays date
today=date.today()
yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')

#Website to get the scores from
webpage_response = requests.get('https://sports.yahoo.com/mlb/scoreboard/?confId=&schedState=2&dateRange=', yesterday)
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the text for the scores, mascots and teams
messy_scores = soup.find_all(attrs={'class':'YahooSans Fw(700)! Va(m) Fz(24px)!'})
messy_odds_mascots = soup.find_all(attrs={'class':'Fw(n) Fz(12px)'})
messy_odds_teams = soup.find_all(attrs={'class':'YahooSans Fw(700)! Fz(14px)!'})
scores = []
odds_teams = []
mascots = []

#Pulls out just the text and stores in list
for a in messy_scores:
    scores.append(a.text)

for a in messy_odds_teams:
    odds_teams.append(a.text)

for a in messy_odds_mascots:
    mascots.append(a.text)                                 

for a in range(0, len(mascots)):
    odds_teams[a] = odds_teams[a] + " " + mascots[a] + " "

odds_both = []

#Creates single list that has scores and teams
for i in range(0, len(scores)):
    odds_both.append([])
    odds_both[i].append(odds_teams[i])
    odds_both[i].append(scores[i])

#Saves the scores in the correct cell based on the associated team
for a in range(0, m):
    for k in range(0, len(odds_both)):
        if ((sheet.cell(a+t, 1).value == odds_both[k][0]) and (sheet.cell(a+t, 9).value is None)):
            for h in range(0, 2):
                sheet.cell(a+t, 9+h).value = int(odds_both[k+h][1])
                sheet.cell(a+t, 11).value = today
                sheet.cell(a+t, 11).number_format = 'dd/mm/yyyy'

#Goes through and deletes rows that don't have a score. Postponed games etc
for a in range(1, m):
    if (sheet.cell(a, 10).value is None):
        sheet.delete_rows(a, 1)

#Finds the first cell that doesn't have a value
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break
m=m-1

#Finds odds for outright wins
#-------------------------------
#Website that has the teams and odds
webpage_response = requests.get('https://www.sportsbet.com.au/betting/baseball/mlb-matches')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the teams and odds
messy_teams = soup.find_all(attrs={'class':'size14_f7opyze Endeavour_fhudrb0 medium_f1wf24vo participantText_fivg86r'})
messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})


#Finds big win and little win odds
#----------------------------------
#Stores website in the web driver
driver.get('https://www.sportsbet.com.au/betting/baseball/mlb-matches')

#Selects the drop down box saying Big Win Little Win
select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
select.select_by_value('Big Win Little Win')

#Waits for 5 seconds to let it load
driver.implicitly_wait(5)

#Pulls out odds and teams
messy_bw_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
messy_bw_teams= driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')


teams = []
odds = []
bw_odds = []
bw_teams = []
bw_both = []

temp2=1
temp3=1
temp4=0

#Gets odds and teams as text and saves in list
for a in messy_teams:
    teams.append(a.get_text())
    
for a in messy_odds:
    odds.append(a.get_text())

for a in messy_bw_odds:
    bw_odds.append(a.text)
    
for a in messy_bw_teams:
    bw_teams.append(a.text)

#Copies values to excel
#----------------------
#Creates list with teams and odds together
for i in range(0, len(bw_teams)//4):
    bw_both.append([])
    temp = bw_teams[i*4].split("Win")
    bw_both[i].append(temp[0])
    for j in range(0, 4):
        bw_both[i].append(bw_odds[temp4+j])
    temp4=temp4+4

temp4=0
count=0

#Copies the values
for i in range (1, len(teams)//2+1):
    for j in range (1, 3):
        temp = teams[i+i-2+j-1].split("(")
        sheet.cell(i+m, j).value = temp[0]
    for k in range (3, 5):
        sheet.cell(i+m, k).value = float(odds[temp2-1+k-3])
    temp2=temp2+6

#Finds difference between big win odds and outright to account for teams that don't have big win yet
diff = ((len(teams)//2)-len(bw_both))

#Copies the big win values
for i in range (1, len(bw_both)+1+diff):
    if(sheet.cell(i+m, 1).value==bw_both[i-1-count][0]):
        for j in range (5, 9):
            sheet.cell(i+m, j).value = float(bw_both[i-1-count][j-4])
    else:
            count=count+1



workbook.save("Baseball_Data.xlsx")
driver.quit()

