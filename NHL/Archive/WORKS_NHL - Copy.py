import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date
import time
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)


filedate=date.today()
filedate = (filedate.strftime('%A %#d %B %Y'))


#Finds the file stored in the location and saves the working sheet and saves a copy

#Personal Laptop
"""
filename = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\NHL\\NHL Tracker.xlsx"
newLocation = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\NHL\\Historical\\NHL Tracker.xlsx"
newName = "NHL " + filedate + ".xlsx"
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\Historical\\NHL Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\Historical\\" + newName)
"""

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\NHL Tracker.xlsx"
#newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\Historical\\NHL Tracker.xlsx"
#newName = "NHL " + filedate + ".xlsx"
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\Historical\\NHL Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\NHL\\Historical\\" + newName)


workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
odds_row=1
    
for odds_row in range(1, 1048576):
    if sheet.cell(odds_row, 9).value is None:
        break
odds_row=odds_row-2

score_row=1
    
for score_row in range(1, 1048576):
    if sheet.cell(score_row, 4).value is None:
        break
score_row=score_row-2

#Finds current round number
prev_round = 0

for previous_round in range(1, 1048576):
    if sheet.cell(previous_round, 1).value is None:
        break

if prev_round != 1:
    round_num = sheet.cell(previous_round-1, 1).value
else:
    round_num = 1

"""
#Pulls out the scores from the previous day
webpage_response = requests.get('https://sports.yahoo.com/nfl/scoreboard/?confId=&dateRange=' + DATE + '&schedState=1')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

messy_teams_scores = soup.find_all(attrs={'class':'Fw(n) Fz(12px)'})
messy_scores = soup.find_all(attrs={'class':'YahooSans Fw(700)! Va(m) Fz(24px)!'})
"""
teams_scores = []
scores = []
difference = odds_row-score_row

for a in range(0, 2*difference):
    teams_scores.append(messy_teams_scores[a].get_text())


for a in range(0, 2*difference):
    scores.append(messy_scores[a].get_text())

for i in range(score_row+1, odds_row+2):
    temp = sheet.cell(i, 2).value.split()
    for j in range(0, len(teams_scores)-1):
        if temp[len(temp)-1] == teams_scores[j]:
            sheet.cell(i, 4).value = int(scores[j])
            sheet.cell(i, 5).value = int(scores[j+1])

#Find out how many games are being played the next day



if difference == 0:

    #Finds odds for outright wins
    #-------------------------------
    #Website that has the teams and odds
    webpage_response = requests.get('https://www.sportsbet.com.au/betting/ice-hockey-us/nhl-matches')
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")
    
    
    #Pulls out the teams and odds
    messy_teams_JUST = soup.find_all(attrs={'class':'size14_f7opyze Endeavour_fhudrb0 medium_f1wf24vo participantText_fivg86r'})
    messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})
    

    #Finds big win and little win odds
    #----------------------------------
    #Stores website in the web driver
    driver.get('https://www.sportsbet.com.au/betting/ice-hockey-us/nhl-matches')

    bets = ["Total Goals", "Money Line", "60 Minute Line", "60 Min Winning Margin", "Both Teams to Score", "60 Minute Draw No Bet", "Will There Be Overtime?", "Highest Scoring Period"]

    messy_more_odds = []
    odds =[]
    count = 5
    row = 2
    teams_JUST = []

    for a in messy_teams_JUST:
        teams_JUST.append(a.get_text())


    #Finds the first row that hasn't been used
    m=1
        
    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break
    m=m-2

    for i in range(0, 2):
        sheet.cell(i+2+m, 2).value = teams_JUST[2*i]
        sheet.cell(i+2+m, 3).value = teams_JUST[2*i+1]
        sheet.cell(i+2+m, 1).value = round_num

    more_odds = []

    for i in range(0, len(bets)):
        print(bets[i])
        row = 2
        teams =[]

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[i] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, 'competition-event-participant')]")
        for a in messy_teams:
            teams.append(a.text)
        
        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []
        odds = []
        messy_total_points = []
        total_points = []
        

        for a in messy_more_odds:
            print(a.text)
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))
        games = len(teams)/3


        if bets[i] == "Total Goals":
            messy_total_points = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, '-outcome-outcome-name')]")
            j = 0
            for a in range(0, len(messy_total_points), 2):
                    temp = messy_total_points[a].text
                    temp2 = temp.split("(")
                    temp3 = temp2[1].split(")")
                    total_points.append(float(temp3[0]))
            
            for row in range(2, 4):
                sheet.cell(row+m, 6).value = total_points[row-2]
                
        if len(more_odds) == games*2:
            j = 0
            for row in range(2, 4):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                j=j+2
            count = count + 2

        elif len(more_odds) == games*3:
            j = 0
            for row in range(2, 4):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                j=j+3
            count = count+3
            
        elif len(more_odds) == games*4:
            j = 0
            for row in range(2, 4):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                j=j+4
            count = count+4
            
        elif len(more_odds) == games*5:
            j = 0
            for row in range(2, 4):
                sheet.cell(row+m, count+2).value = more_odds[j]
                sheet.cell(row+m, count+1+2).value = more_odds[j+1]
                sheet.cell(row+m, count+2+2).value = more_odds[j+2]
                sheet.cell(row+m, count+3+2).value = more_odds[j+3]
                sheet.cell(row+m, count+4+2).value = more_odds[j+4]
                j=j+5
            count = count+5
        print("Done")

workbook.save("NHL Tracker.xlsx")

