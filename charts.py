"""
*** TO DO ***

Remove legend

Last XXX number of days

Fix up the bets that have different counts

Make the one file so can just click the python button

Clean up what isn't needed

"""



import openpyxl as xl
from openpyxl.chart import (LineChart, Reference, Series)
from string import ascii_uppercase

#Opens and creates the relevant spreadsheets
#data ="C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Baseball_Data.xlsx"
data ="C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Baseball_Data.xlsx"
data_book = xl.load_workbook(data, data_only=True)
new_book = xl.Workbook()
new_book_data = new_book.worksheets[0]
new_book_data.title = "Data"
new_book_charts = new_book.create_sheet()
new_book_charts.title = "Charts"
analysis_sheet = data_book.worksheets[2]
summary_sheet = data_book.worksheets[0]

categories = []
i=2


#Gets all of the categories of bets
while ((analysis_sheet.cell(1, i).value is None) == False):
    categories.append(analysis_sheet.cell(1, i).value)
    i = i+1

new_date = summary_sheet.cell(2, 11).value
c=0
#Writes all of the categories of best across the top row of new spreadsheet
for a in range(2, 2*len(categories)+2, 2):
    new_book_data.cell(1, a+1).value = categories[c]
    c=c+1

               
j=2
k = 2
f = 0
running_average_sum = 0.0
running_average = 0.0
total_count = 0
count = 0
multiplier_sum = 0
total_day = 0
day_array = [0] * (len(categories)+1)


#Loops through for each of the different days
while (new_date is None) == False:


    count = 0
    #Finds the count of the number of games that day
    while (summary_sheet.cell(j, 11).value == new_date):
        count = count+1
        j = j+1
    day = 0
    multiplier = 0
    total_count = total_count + count

    #Works out what the multiplier is for that day
    for b in range(1, len(categories)+1):
        for a in range(j-count, j):
            day = day + analysis_sheet.cell(a+5, b+1).value
            multiplier = day/count

        #Stores the values of the multiplier and the date
        day_array[b] = day_array[b] + day
        new_book_data.cell(k, 1).value = new_date
        new_book_data.cell(k, 1).number_format = 'dd/mm/yyyy'
        new_book_data.cell(k, 2).value = 1
        new_book_data.cell(k, 2*b+1).value = multiplier
        new_book_data.cell(k, 2*b+2).value = multiplier

        #Can probably delete this
        for t in range(2, k+1):
            multiplier_sum = multiplier_sum + new_book_data.cell(t, 2*b+1).value
            running_average_sum = running_average_sum + new_book_data.cell(t, 2*b+2).value

        #Calculates the running average
        running_average = day_array[b]/total_count
        new_book_data.cell(k, 2*b+2).value = running_average
        day = 0
        multiplier = 0
        running_average = 0
        running_average_sum = 0
    
    k = k + 1
    f = f + 1
    new_date = summary_sheet.cell(j, 11).value


#Widens the column and hides columns
new_book_data.column_dimensions['A'].width = 12
#new_book_data.column_dimensions['B'].hidden = True

#Creates the charts
for a in range(1, len(categories)+1):

    chart = LineChart()
    chart.title = categories[a-1]
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data = Reference(new_book_data, min_col = 2*a+1, min_row = 2, max_col = 2*a+1, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "B"+str(a*18-16))

    chart = LineChart()
    chart.title = categories[a-1]
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data2 = Reference(new_book_data, min_col = 2*a+2, min_row = 2, max_col = 2*a+2, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data2)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "K"+str(a*18-16))

new_book.active = new_book["Charts"]
new_book.save("Baseball_Data_Charts.xlsx")

