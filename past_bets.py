"""
open file and check the bets from previous week

read from the Disposals tracking file
pull out all of the odds from this and save it new tab of the past_bets doc
title it round X

Has a section for each game and split into multi and singles
multi
    highest odds difference and above 90% percentage
singles
    5 highest above 60% percentage
"""

    
import openpyxl as xl

disposals_filename ="C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Disposals Tracking.xlsx"
disposals = xl.load_workbook(disposals_filename)


bets_filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Bets.xlsx"
bets = xl.Workbook(bets_filename)
bets_sheet = bets.create_sheet(0)
bets_sheet.title = "Jay"
bets_sheet = bets.create_sheet(0)
bets_sheet.title = "Blah"
bets_sheet = bets.create_sheet(0)
bets_sheet.title = "Blahhhh"

bets.save("Bets.xlsx")


