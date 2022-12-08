"""
Next Steps:
Fix it so only tomorrow's games count
    Currently pulls out today's date in same format
    Get it to compare to all of the dates
    If tomorrow's date matches then run the script to pull the odds

Sort out the past games score check

Comments

"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = "C:\\Users\\jhansen3\\AppData\\Local\\Google\\Chrome Beta\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(service=Service(ChromeDriverManager(version='104.0.5112.20').install()), options = chrome_options)


filedate=date.today()
filedate = (filedate.strftime('%A %#d %B %Y'))

print(filedate)

"""
#Finds the file stored in the location and saves the working sheet and saves a copy
filename ="C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Baseball_Data.xlsx"
newLocation = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\Baseball_Data.xlsx"
newName = "Baseball_Data_" + filedate + ".xlsx"
shutil.copyfile(filename, newLocation)
os.rename("C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\Baseball_Data.xlsx", "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Historical\\" + newName)
workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]
"""

"""
#Gets the scores for the previous day
#------------------------------------

#Finds the first cell without a score
t=1
    
for t in range(1, 1048576):
    if sheet.cell(t, 9).value is None:
        break
t=t

#Gets yesterdays date
today=date.today()
yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')

#Website to get the scores from
webpage_response = requests.get('https://sports.yahoo.com/mlb/scoreboard/?confId=&schedState=2&dateRange=', yesterday)
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the text for the scores, mascots and teams
messy_scores = soup.find_all(attrs={'class':'YahooSans Fw(700)! Va(m) Fz(24px)!'})
messy_odds_mascots = soup.find_all(attrs={'class':'Fw(n) Fz(12px)'})
messy_odds_teams = soup.find_all(attrs={'class':'YahooSans Fw(700)! Fz(14px)!'})
scores = []
odds_teams = []
mascots = []

#Pulls out just the text and stores in list
for a in messy_scores:
    scores.append(a.text)

for a in messy_odds_teams:
    odds_teams.append(a.text)

for a in messy_odds_mascots:
    mascots.append(a.text)                                 

for a in range(0, len(mascots)):
    odds_teams[a] = odds_teams[a] + " " + mascots[a] + " "

odds_both = []

#Creates single list that has scores and teams
for i in range(0, len(scores)):
    odds_both.append([])
    odds_both[i].append(odds_teams[i])
    odds_both[i].append(scores[i])

#Saves the scores in the correct cell based on the associated team
for a in range(0, m):
    for k in range(0, len(odds_both)):
        if ((sheet.cell(a+t, 1).value == odds_both[k][0]) and (sheet.cell(a+t, 9).value is None)):
            for h in range(0, 2):
                sheet.cell(a+t, 9+h).value = int(odds_both[k+h][1])
                sheet.cell(a+t, 11).value = today
                sheet.cell(a+t, 11).number_format = 'dd/mm/yyyy'

#Goes through and deletes rows that don't have a score. Postponed games etc
for a in range(1, m):
    if (sheet.cell(a, 10).value is None):
        sheet.delete_rows(a, 1)

#Finds the first cell that doesn't have a value
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break
m=m-1
"""

"""
#Finds the next day with games and the number of games
#------------------------------------------------------
webpage_response = requests.get('https://www.theguardian.com/football/premierleague/fixtures')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

next_date_web_element = soup.find_all(attrs={'class':'football-matches__day'})

for i in range(0, len(next_date_web_element)):
    next_date = next_date_web_element[i].find_all(attrs={'class':'date-divider'})
    print(next_date[0].get_text())

next_date = next_date_web_element[0].find_all(attrs={'class':'date-divider'})

count_games = len(next_date_web_element[0].find_all(attrs={'class':'football-teams__battleline'}))

"""


#Finds odds for outright wins
#-------------------------------
#Website that has the teams and odds
webpage_response = requests.get('https://www.sportsbet.com.au/betting/soccer/germany/german-bundesliga')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the teams and odds
messy_teams = soup.find_all(attrs={'class':'size12_fq5j3k2 normal_fgzdi7m caption_f4zed5e'})
messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})







#Finds big win and little win odds
#----------------------------------
#Stores website in the web driver
driver.get('https://www.sportsbet.com.au/betting/soccer/germany/german-bundesliga')

bets = ["Over/Under 2.5 Goals", "Both Teams To Score", "Draw No Bet", "Double Chance"]

messy_more_odds = []
teams =[]
odds =[]
count = 6
row = 2

for a in messy_teams:
    teams.append(a.get_text())
    
for a in messy_odds:
    odds.append(a.get_text())

filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Bundesliga Tracker.xlsx"
workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break
m=m-2

#ADD COUNT OF GAMES HERE

for i in range(0, 10):
    sheet.cell(i+2+m, 1).value = teams[3*i]
    sheet.cell(i+2+m, 2).value = teams[3*i+2]
    sheet.cell(i+2+m, 3).value = float(odds[3*i])
    sheet.cell(i+2+m, 4).value = float(odds[3*i+1])
    sheet.cell(i+2+m, 5).value = float(odds[3*i+2])


for i in range(0, len(bets)):
    print(bets[i])
    row = 2
    more_odds = []
    select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
    select.select_by_value(bets[i])

    driver.implicitly_wait(5)
    
    messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    odds = []

    for a in messy_more_odds:
        more_odds.append(float(a.text))


    if len(more_odds) != len(teams):
        for j in range(0, 10):
            sheet.cell(row+m, count).value = more_odds[j]
            sheet.cell(row+m, count+1).value = more_odds[j+1]
            j=j+2
            row = row +1
        count = count + 2
        
    else:
        for j in range(0, 10):
            sheet.cell(row+m, count).value = more_odds[j]
            sheet.cell(row+m, count+1).value = more_odds[j+1]
            sheet.cell(row+m, count+2).value = more_odds[j+2]
            j=j+3
            row = row+1
        count = count+3


workbook.save("Bundesliga Tracker.xlsx")
