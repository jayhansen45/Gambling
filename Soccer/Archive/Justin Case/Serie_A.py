"""
Next Steps:

Add some shit about the next game?
    Email self when the next game is so can run before
    Set reminder would be even better

Comments

"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = "C:\\Users\\jhansen3\\AppData\\Local\\Google\\Chrome Beta\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(service=Service(ChromeDriverManager(version='104.0.5112.20').install()), options = chrome_options)


filedate=date.today()
filedate = (filedate.strftime('%A %#d %B %Y'))


#Below probably isn't needed tbh
"""
#Finds the next day with games and the number of games
#------------------------------------------------------
webpage_response = requests.get('https://www.theguardian.com/football/premierleague/fixtures')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

next_date_web_element = soup.find_all(attrs={'class':'football-matches__day'})

for i in range(0, len(next_date_web_element)):
    next_date = next_date_web_element[i].find_all(attrs={'class':'date-divider'})
    print(next_date[0].get_text())

next_date = next_date_web_element[0].find_all(attrs={'class':'date-divider'})

count_games = len(next_date_web_element[0].find_all(attrs={'class':'football-teams__battleline'}))

"""

#Finds the file stored in the location and saves the working sheet and saves a copy

filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Serie A Tracker.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Historical\\Serie A Tracker.xlsx"
newName = "Serie A Tracker " + filedate + ".xlsx"
shutil.copyfile(filename, newLocation)
os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Historical\\Serie A Tracker.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\Historical\\" + newName)
workbook = xl.load_workbook(filename)
sheet = workbook.worksheets[0]

#Finds the first row that hasn't been used
odds_row=1
    
for odds_row in range(1, 1048576):
    if sheet.cell(odds_row, 1).value is None:
        break
odds_row=odds_row-2

score_row=1
    
for score_row in range(1, 1048576):
    if sheet.cell(score_row, 15).value is None:
        break
score_row=score_row-2

#Finds current round number
prev_round = 1

for previous_round in range(1, 1048576):
    if sheet.cell(previous_round, 17).value is None:
        break

round_num = sheet.cell(previous_round-1, 17).value+1

webpage_response = requests.get('https://www.theguardian.com/football/serieafootball/results')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

messy_teams_scores = soup.find_all(attrs={'class':'team-name__long'})
messy_scores = soup.find_all(attrs={'class':'football-team__score'})

teams_scores = []
scores = []
difference = odds_row-score_row

for a in range(0, 2*difference):
    if messy_teams_scores[a].get_text() == "Monza":
        teams_scores.append("AC Monza")
    elif messy_teams_scores[a].get_text() == "Cremonese":
        teams_scores.append("US Cremonese")
    elif messy_teams_scores[a].get_text() == "Spurs":
        teams_scores.append("Tottenham")
    else:    
        teams_scores.append(messy_teams_scores[a].get_text())

for a in range(0, 2*difference):
    scores.append(messy_scores[a].get_text())


for i in range(score_row+1, odds_row+2):
    for j in range(0, len(teams_scores)):
        if sheet.cell(i, 1).value == teams_scores[j]:
            sheet.cell(i, 15).value = int(scores[j])
            sheet.cell(i, 16).value = int(scores[j+1])



#Finds odds for outright wins
#-------------------------------
#Website that has the teams and odds
webpage_response = requests.get('https://www.sportsbet.com.au/betting/soccer/italy/italian-serie-a')
webpage = webpage_response.content
soup = BeautifulSoup(webpage, "html.parser")

#Pulls out the teams and odds
messy_teams = soup.find_all(attrs={'class':'size12_fq5j3k2 normal_fgzdi7m caption_f4zed5e'})
messy_odds = soup.find_all(attrs={'class':'size14_f7opyze bold_f1au7gae priceTextSize_frw9zm9'})


#Finds big win and little win odds
#----------------------------------
#Stores website in the web driver
driver.get('https://www.sportsbet.com.au/betting/soccer/italy/italian-serie-a')

bets = ["Over/Under 2.5 Goals", "Both Teams To Score", "Draw No Bet", "Double Chance"]

messy_more_odds = []
teams =[]
odds =[]
count = 6
row = 2

for a in messy_teams:
    teams.append(a.get_text())
    
for a in messy_odds:
    odds.append(a.get_text())



#Finds the first row that hasn't been used
m=1
    
for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break
m=m-2

for i in range(0, 10):
    sheet.cell(i+2+m, 1).value = teams[3*i]
    sheet.cell(i+2+m, 2).value = teams[3*i+2]
    sheet.cell(i+2+m, 3).value = float(odds[3*i])
    sheet.cell(i+2+m, 4).value = float(odds[3*i+1])
    sheet.cell(i+2+m, 5).value = float(odds[3*i+2])
    sheet.cell(i+2+m, 17).value = round_num


more_odds = []

print(teams)

for i in range(0, len(bets)):
    print(bets[i])
    row = 2

    select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
    select.select_by_value(bets[i])

    driver.implicitly_wait(5)
    
    messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    more_odds = []
    odds = []

    for a in messy_more_odds:
        if a.text == "Suspended":
            more_odds.append(0)
        else:
            more_odds.append(float(a.text))

    print(more_odds)

    if len(more_odds) != len(teams):
        j = 0
        for row in range(2, 12):
            sheet.cell(row+m, count).value = more_odds[j]
            sheet.cell(row+m, count+1).value = more_odds[j+1]
            j=j+2
        count = count + 2
        
    else:
        j = 0
        for row in range(2, 12):
            sheet.cell(row+m, count).value = more_odds[j]
            sheet.cell(row+m, count+1).value = more_odds[j+1]
            sheet.cell(row+m, count+2).value = more_odds[j+2]
            j=j+3
        count = count+3



workbook.save("Serie A Tracker.xlsx")

