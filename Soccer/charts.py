"""
*** TO DO ***

Remove legend

Last XXX number of days

Fix up the bets that have different counts

Make the one file so can just click the python button

Clean up what isn't needed

"""



import openpyxl as xl
from openpyxl.chart import (LineChart, Reference, Series)
from string import ascii_uppercase

#Opens and creates the relevant spreadsheets
data ="C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Soccer\\EPL Tracker.xlsx"
#data ="C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Baseball_Data.xlsx"
data_book = xl.load_workbook(data, data_only=True)
new_book = xl.Workbook()
new_book_data = new_book.worksheets[0]
new_book_data.title = "Data"
new_book_charts = new_book.create_sheet()
new_book_charts.title = "Charts"
analysis_sheet = data_book.worksheets[2]
summary_sheet = data_book.worksheets[0]

categories = []
i=3


#Gets all of the categories of bets
for i in range(1, 1000):
    if (analysis_sheet.cell(1, i).value != "Home Team is Fave") and (analysis_sheet.cell(1, i).value is None) == False:
        categories.append(analysis_sheet.cell(1, i).value)
new_round = analysis_sheet.cell(7, 1).value
c=0
#Writes all of the categories of bets across the top row of new spreadsheet
for a in range(2, 2*len(categories)+2, 2):
    new_book_data.cell(1, a+1).value = categories[c]
    c=c+1


#Loops through for each of the different days
round_row = 7
total_round = 0
round_num = 1
k = 2
for category in range(2, 2*len(categories)+3):
    round_row = 7
    new_round = analysis_sheet.cell(7, 1).value
    running_multiplier = 0
    running_bets = 0
    running_sum = 0
    k = 2

    while analysis_sheet.cell(round_row, 1).value != 0:
        total_round = 0
        day_bets = 0
        wins = 0
        multiplier = 0


        while new_round == analysis_sheet.cell(round_row, 1).value:
            round_row = round_row +1
            total_round = total_round +1

        if (analysis_sheet.cell(1, category).value is None) == True:
            for i in range(0, total_round):
                day_bets = day_bets + analysis_sheet.cell(round_row - total_round + i, category).value
                wins = wins + analysis_sheet.cell(round_row - total_round + i, category + 1).value
                
            running_sum = running_sum + wins
            running_bets = running_bets + day_bets
            
            if day_bets >0:
                multiplier = wins/day_bets

            else:
                multiplier = 0

            running_multiplier = running_sum/running_bets
        
            new_book_data.cell(k, 1).value = new_round
            new_book_data.cell(k, 2).value = 1     
            new_book_data.cell(k, category).value = multiplier
            new_book_data.cell(k, category+1).value = running_multiplier

        new_round = analysis_sheet.cell(round_row, 1).value
        k = k + 1

print(k)        
#Widens the column and hides columns
new_book_data.column_dimensions['A'].width = 12
#new_book_data.column_dimensions['B'].hidden = True


#Creates the charts
for a in range(1, len(categories)+1):

    chart = LineChart()
    chart.title = categories[a-1]
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data = Reference(new_book_data, min_col = 2*a+1, min_row = 2, max_col = 2*a+1, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "B"+str(a*18-16))

    chart = LineChart()
    chart.title = categories[a-1]
    chart.y_axis.title = 'Multiplier'
    chart.x_axis.title = "Dates"

    data2 = Reference(new_book_data, min_col = 2*a+2, min_row = 2, max_col = 2*a+2, max_row = k-1)
    dates = Reference(new_book_data, min_col = 1, min_row = 2, max_col = 1, max_row = k-1)
    ones = Reference(new_book_data, min_col = 2, min_row = 2, max_col = 2, max_row = k-1)
    series = Series(ones)
    chart.legend.visible = False
    chart.add_data(data2)
    chart.series.append(series)
    chart.set_categories(dates)
    new_book_charts.add_chart(chart, "K"+str(a*18-16))

new_book.active = new_book["Charts"]
new_book.save("EPL Charts.xlsx")

