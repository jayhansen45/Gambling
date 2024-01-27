import requests
import openpyxl as xl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import time
import shutil
import os
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
#chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(executable_path=r"C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Other\\Chrome Driver\\chromedriver.exe", options = chrome_options)
#driver = webdriver.Chrome(ChromeDriverManager(version='118.0.5993.70').install(), options = chrome_options)

#date
filedate=date.today()+timedelta(days=2)
day = (filedate.strftime('%m-%d-%Y'))

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\FFlyer\\Data.xlsx"

workbook = xl.load_workbook(filename)
sheet = workbook['Data']



#site = "https://book.virginaustralia.com/dx/VADX/#/flight-selection?journeyType=one-way&activeMonth="+day+"&locale=en-GB&awardBooking=false&class=First&ADT=1&CHD=0&INF=0&origin=MEL&destination=ATH&date="+day+"&promoCode=&execution=undefined"
site = "https://book.virginaustralia.com/dx/VADX/#/date-selection?journeyType=one-way&activeMonth="+day+"&locale=en-GB&awardBooking=false&searchType=BRANDED&class=First&ADT=1&CHD=0&INF=0&origin=MEL&destination=ABE&promoCode=&direction=0&execution=undefined"
driver.get(site)
time.sleep(3)

#For loop goes here
for c in range(19, 100):
    #No flights catch
    try:
        time.sleep(3)
        print("Are there flights?")
        flights = driver.find_element(By.XPATH, '//*[@data-translation="flightNotFound.title"]')
    except:
        print("Yes")
        #Stores country
        departing = driver.find_element(By.XPATH, '//*[@data-translation="summaryBar.vaDescription.fullOneWay"]')

        airport = (departing.text).split("Melbourne to ")
        final_destination = airport[1].split(", departing")
        
        print(final_destination[0])
        time.sleep(3)

        for b in range(0, 2):
            try:
                print("What denonimation?")
                status = driver.find_element(By.XPATH, '//*[@data-translation="app.currency.FFCURRENCY.symbol"]')
                print("In Points")
                print("Sort options")
                #Changes sort options
                try:
                    print("Is there more than one?")
                    clickable = driver.find_element(By.XPATH, '//*[@id="dxp-sort-toggle-button"]')
                    clickable.click()
                    time.sleep(3)
                    print("Yes")
                    print("Sort by price")
                    clickable = driver.find_element(By.XPATH, '//*[@id="radio-flight-sort-0-price-asc"]')
                    clickable.click()
                    time.sleep(3)
                    clickable = driver.find_element(By.XPATH, '//*[@class="dxp-button va-modal-action-button spark-btn update action-primary secondary medium"]')
                    clickable.click()
                    time.sleep(3)

                #gets the points
                except:
                    print("No, only one")
                    print("Get points")
                    points = driver.find_elements(By.XPATH, '//*[@class="number"]')

                    print("Switching to currency")
                    clickable = driver.find_element(By.XPATH, '//*[@id="flight-selection-points-currency-toggle-0"]')
                    clickable.click()
                    time.sleep(3)
                else:
                    print("Get points")
                    points = driver.find_elements(By.XPATH, '//*[@class="number"]')

                    print("Switching to currency")
                    clickable = driver.find_element(By.XPATH, '//*[@id="flight-selection-points-currency-toggle-0"]')
                    clickable.click()
                    time.sleep(3)

            except:
                print("In Currency")
                print("Sort options")
                #Changes sort options
                try:
                    print("Is there more than one?")
                    clickable = driver.find_element(By.XPATH, '//*[@id="dxp-sort-toggle-button"]')
                    clickable.click()
                    time.sleep(3)
                    print("Yes")
                    print("Sort by price")
                    clickable = driver.find_element(By.XPATH, '//*[@id="radio-flight-sort-0-price-asc"]')
                    clickable.click()
                    time.sleep(3)
                    clickable = driver.find_element(By.XPATH, '//*[@class="dxp-button va-modal-action-button spark-btn update action-primary secondary medium"]')
                    clickable.click()
                    time.sleep(3)

                #Gets dollars
                except:
                    print("No, only one")
                    print("Get Dollars")
                    dollars = driver.find_elements(By.XPATH, '//*[@class="number"]')
                    dollars = dollars[10].text
                    time.sleep(3)
                    
                    print("Switching to points")
                    clickable = driver.find_element(By.XPATH, '//*[@id="flight-selection-points-currency-toggle-1"]')
                    clickable.click()
                    time.sleep(3)
                else:
                    print("Get Dollars")
                    dollars = driver.find_elements(By.XPATH, '//*[@class="number"]')
                    dollars = dollars[10].text
                    time.sleep(3)
                    
                    print("Switching to points")
                    clickable = driver.find_element(By.XPATH, '//*[@id="flight-selection-points-currency-toggle-1"]')
                    clickable.click()
                    time.sleep(3)



        #Copies values into excel
        i = 0

        #Finds first row that hasn't been used
        m=3

        for m in range(3, 1048576):
            if sheet.cell(m, 2).value is None:
                break
            
        print(final_destination[0])
        for a in points:
            if a.text!="" or a.text!="0" or a.text!=" ":
                sheet.cell(m+i, 3).value = day
                sheet.cell(m+i, 2).value = final_destination[0]
                sheet.cell(m+i, 6).value = dollars
                if a.text.find('.') == -1:
                    sheet.cell(m+i, 4).value = a.text
                else:
                    i = i-1
                    sheet.cell(m+i, 5).value = a.text
                i = i+1


    print("Next country")
    print(" ")
    #Clicks edit search button
    clickable = driver.find_element(By.XPATH, '//*[@class="dxp-button va-edit-search-button is-unstyled"]')
    clickable.click()
    time.sleep(3)

    #Opens drop down box
    clickable = driver.find_element(By.XPATH, '//*[@id="arriving-airport0"]')
    clickable.click()
    time.sleep(3)

    #Goes to the next airport
    #clickable = driver.find_element(By.XPATH, '//*[@id="react-autowhatever-arriving-airport0-auto-suggest--item-2"]')
    search = '//*[@id="react-autowhatever-arriving-airport0-auto-suggest--item-'+ str(c) + '"]'
    clickable = driver.find_element(By.XPATH, search)
    clickable.click()
    time.sleep(3)
    clickable = driver.find_elements(By.XPATH, '//*[@id="dxp-page-navigation-continue-button"]')
    if len(clickable) == 1:
        clickable[0].click()
    else:
        clickable[1].click()
    time.sleep(3)

    clickable = driver.find_element(By.XPATH, '//*[@class="dxp-button va-modal-action-button action-primary secondary medium"]')
    clickable.click()
    time.sleep(3)
    workbook.save("Data.xlsx")


driver.quit()
workbook.save("Data.xlsx")


