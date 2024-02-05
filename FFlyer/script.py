"""

"""


import requests
import openpyxl as xl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import time
import datetime
import shutil
import math
import os
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument("--start-maximized")
#chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(executable_path=r"C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Other\\Chrome Driver\\chromedriver.exe", options = chrome_options)
#driver = webdriver.Chrome(ChromeDriverManager(version='118.0.5993.70').install(), options = chrome_options)



#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\FFlyer\\Data.xlsx"

workbook = xl.load_workbook(filename)
sheet = workbook['Data']


start = 0
start_date = 57
#57 = April 20
iterator = 0

airports = ["EBB", "DAR", "FUK", "HND", "HNL", "HRE", "ITM", "JNB", "JRO", "KIX", "KMI", "KOA", "LUN", "MLE", "OGG", "SEZ", "ZNZ", "CCK", "CMB", "MPM", "XCH", "HIJ", "OIT", "OKA", "VLI", "KTM", "KKJ", "TTJ", "NRT", "LIM"]

start_time = time.time()

for q in range(20+start_date, 200):
    #date
    filedate=date.today()+timedelta(days=q)
    day = (filedate.strftime('%m-%d-%Y'))
    #For loop goes here
    for c in range(0+start, len(airports)):
        site = "https://book.virginaustralia.com/dx/VADX/#/flight-selection?journeyType=one-way&activeMonth="+day+"&direction=0&locale=en-GB&awardBooking=true&origin=MEL&destination="+airports[c]+"&class=First&ADT=1&CHD=0&INF=0&date="+day+"&promoCode=&execution=undefined"
        driver.get(site)
        time.sleep(3)
        driver.refresh()
        time.sleep(3)
        driver.execute_script("document.body.style.zoom='67%'")
        time.sleep(3)
        check = 0

        departing = driver.find_element(By.XPATH, '//*[@data-translation="summaryBar.vaDescription.fullOneWay"]')

        airport = (departing.text).split("Melbourne to ")
        final_destination = airport[1].split(", departing")
        
        print(final_destination[0])
        
        try:
            time.sleep(3)
            print("Are there flights?")
            flights = driver.find_element(By.XPATH, '//*[@data-translation="flightNotFound.title"]')
        except:
            print("Yes")
            #Stores country
            departing = driver.find_element(By.XPATH, '//*[@data-translation="summaryBar.vaDescription.fullOneWay"]')

            airport = (departing.text).split("Melbourne to ")
            final_destination = airport[1].split(", departing")
            
            try:
                clickable = driver.find_element(By.XPATH, '//*[@id="dxp-sort-toggle-button"]')
                clickable.click()
                time.sleep(3)
                print("Sort by price")
                clickable = driver.find_element(By.XPATH, '//*[@id="radio-flight-sort-0-price-asc"]')
                clickable.click()
                time.sleep(3)
                clickable = driver.find_element(By.XPATH, '//*[@class="dxp-button va-modal-action-button spark-btn update action-primary secondary medium"]')
                clickable.click()
                time.sleep(3)
                print("Get points")
                points = driver.find_elements(By.XPATH, '//*[@class="number"]')
                time.sleep(3)
            except:
                print("Get points")
                points = driver.find_elements(By.XPATH, '//*[@class="number"]')
                time.sleep(3)

            if check == 0:
                #Copies values into excel
                i = 0

                #Finds first row that hasn't been used
                m=3

                for m in range(3, 1048576):
                    if sheet.cell(m, 2).value is None:
                        break

                
                for a in points:
                    if a.text!="" and a.text!="0" and a.text!=" ":
                        sheet.cell(m+i, 2).value = final_destination[0]
                        sheet.cell(m+i, 3).value = day
                        if a.text.find('.') == -1:
                            sheet.cell(m+i, 4).value = a.text
                        else:
                            i = i-1
                            sheet.cell(m+i, 5).value = a.text
                        i = i+1

        else:
            print("No")
        split = time.time()

        sec = round(split-start_time, 0)

        convert = str(datetime.timedelta(seconds = sec))
        print(convert, "elapsed")
        print("Next Country")
        print()
        workbook.save("Data.xlsx")
        start = 0
        if iterator == 30:
            print("Closing driver")
            driver.close()
            driver.quit()
            time.sleep(3)

            driver = webdriver.Chrome(executable_path=r"C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Other\\Chrome Driver\\chromedriver.exe", options = chrome_options)
            iterator = 0
            print()
        else:
            iterator = iterator + 1


workbook.save("Data.xlsx")


