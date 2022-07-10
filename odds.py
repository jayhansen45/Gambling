"""
NEXT STEPS:
Add a "STEPS" Thing below

Fix the games issue
    Scoot across the top row starting at 2 and when != avg it is the number of games


"""

import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import warnings
import shutil
import os

urls = []
games = 15
url = ""
warnings.filterwarnings("ignore")
home_array = []
away_array = []

while url != "Run Please":
    url = input("Enter the URL of the game: ")
    if url != "Run Please":
        urls.append(url)

j=0

for j in range(0, len(urls)):
    #Initiates all of the arrays
    players_15 = []
    odds_15 = []
    players_20 = []
    odds_20 = []
    players_25 = []
    odds_25 = []
    players_30 = []
    odds_30 = []
    players_35 = []
    odds_35 = []

    #Chomedriver options for some reason
    chrome_options = webdriver.ChromeOptions()
    chrome_options.binary_location = "C:\\Users\\jhansen3\\AppData\\Local\\Google\\Chrome Beta\\Application\\chrome.exe"
    chrome_options.headless = True
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--incognito")

    os.chmod('C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\chromedriver_win32\\chromedriver.exe', 0o755)


    #Switch the below comments when swapping laptop
    filename ="C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Disposals Tracking.xlsx"
    #filename = "C:\\Users\\jayha\\Documents\\Gambling\\Automated\\Disposals Tracking.xlsx"

    #driver = webdriver.Chrome(executable_path=r"C:\Users\jhansen3\OneDrive - KPMG\Documents\Python\Gambling\chromedriver_win32\chromedriver.exe")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager(version='104.0.5112.20').install()), options=chrome_options)


    #Driver to get the info from that specific URL
    driver.get(urls[j])

    home_team = driver.find_elements(By.XPATH, '//*[@data-automation-id="event-participant-1"]')
    away_team = driver.find_elements(By.XPATH, '//*[@data-automation-id="event-participant-2"]')

    home_array.append(home_team[0].text)
    away_array.append(away_team[0].text)
    

    #Finds and clicks the Top Markets then the Disposal Markets
    element = driver.find_elements(By.XPATH, '//*[@data-automation-id="market-group-accordion-header-title"]')

    for i in range(0, len(element)):
        if element[i].text == "Top Markets":
            element[i].click()

    elements = driver.find_elements(By.XPATH, '//*[@data-automation-id="market-group-accordion-header"]')

    for i in range(0, len(elements)):
        if elements[i].text == "Disposal Markets":
            elements[i].click()


    #Finds and clicks each of the disposal markets
    disposals = driver.find_elements(By.XPATH, '//*[@data-automation-id="accordion-header"]')

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 15 or More Disposals":
            disposals[i].click()

    #Pulls the data into an array
    messy_players_15 = driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')
    messy_odds_15 = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')


    #Appends the text to a new array
    for i in range(0, len(messy_players_15)):
        players_15.append(messy_players_15[i].text)
        if messy_odds_15[i].text != "SUS":
            odds_15.append(messy_odds_15[i].text)
        else:
            odds_15.append(0)

    #Closes the accordion heading
    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 15 or More Disposals":
            disposals[i].click()


    #As above
    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 20 or More Disposals":
            disposals[i].click()


    messy_players_20 = driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')
    messy_odds_20 = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    for i in range(0, len(messy_players_20)):
        players_20.append(messy_players_20[i].text)
        if messy_odds_20[i].text != "SUS":
            odds_20.append(messy_odds_20[i].text)
        else:
            odds_20.append(0)

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 20 or More Disposals":
            disposals[i].click()

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 25 or More Disposals":
            disposals[i].click()

    messy_players_25 = driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')
    messy_odds_25 = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    for i in range(0, len(messy_players_25)):
        players_25.append(messy_players_25[i].text)
        if messy_odds_25[i].text != "SUS":
            odds_25.append(messy_odds_25[i].text)
        else:
            odds_25.append(0)

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 25 or More Disposals":
            disposals[i].click()


    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 30 or More Disposals":
            disposals[i].click()

    messy_players_30 = driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')
    messy_odds_30 = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    for i in range(0, len(messy_players_30)):
        players_30.append(messy_players_30[i].text)
        if messy_odds_30[i].text != "SUS":
            odds_30.append(messy_odds_30[i].text)
        else:
            odds_30.append(0)

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 30 or More Disposals":
            disposals[i].click()

    for i in range(0, len(disposals)):
        if disposals[i].text == "To Get 35 or More Disposals":
            disposals[i].click()

    messy_players_35 = driver.find_elements(By.XPATH, '//*[contains(@data-automation-id, "-column-grid-outcome-name")]')
    messy_odds_35 = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')

    for i in range(0, len(messy_players_35)):
        players_35.append(messy_players_35[i].text)
        if messy_odds_35[i].text != "SUS":
            odds_35.append(messy_odds_35[i].text)
        else:
            odds_35.append(0)

    
    driver.quit()

    #Stores the data in the Disposals Tracking spreadsheet
    workbook = xl.load_workbook(filename)
    sheet = workbook.worksheets[0]

    i = 1
    
    #Finds the bottom cell
    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break
    m=m-1

    #Loops through and stores the value in the correct cell
    for i in range(0, len(players_15)):
        for a in range(1, m):
            if players_15[i] == sheet.cell(a, 2).value:
                sheet.cell(a, games+7).value = float(odds_15[i])

    for i in range(0, len(players_20)):
        for a in range(1, m):
            if players_20[i] == sheet.cell(a, 2).value:
                sheet.cell(a, games+11).value = float(odds_20[i])


    for i in range(0, len(players_25)):
        for a in range(1, m):
            if players_25[i] == sheet.cell(a, 2).value:
                sheet.cell(a, games+15).value = float(odds_25[i])


    for i in range(0, len(players_30)):
        for a in range(1, m):
            if players_30[i] == sheet.cell(a, 2).value:
                sheet.cell(a, games+19).value = float(odds_30[i])

    for i in range(0, len(players_35)):
        for a in range(1, m):
            if players_35[i] == sheet.cell(a, 2).value:
                sheet.cell(a, games+23).value = float(odds_35[i])
    j = j+1

    workbook.save("Disposals Tracking.xlsx")



bets_filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Bets.xlsx"
bets = xl.load_workbook(bets_filename)
sheets = bets.sheetnames

for p in range(0, len(home_array)):

    y=0

    if sheets[0] != ("Round " + str(games+1)):
        bets_sheet = bets.create_sheet("Round " + str(games+1), 0)
    else:
        bets_sheet = bets.worksheets[0]
        y=1048576
        while bets_sheet.cell(y, 2).value == None:
            y=y-1
        y=y+2


    bets_sheet.column_dimensions['B'].width = 15
    bets_sheet.column_dimensions['C'].width = 15
    bets_sheet.column_dimensions['D'].width = 15
    bets_sheet.column_dimensions['E'].width = 20
    bets_sheet.column_dimensions['F'].width = 15




    bets_sheet.cell(2+y, 2).value = home_array[p] + " Vs " + away_array[p]
    bets_sheet.cell(2+y, 2).font = Font(bold = True)
    bets_sheet.cell(2+y, 2).fill = PatternFill("solid", fgColor = "BDD7EE")
    bets_sheet.merge_cells('B' + str(2+y) + ':F' + str(2+y))
    bets_sheet.cell(2+y, 2).alignment = Alignment(horizontal = "center")

    bets_sheet.cell(3+y, 2).value = "Player"
    bets_sheet.cell(3+y, 3).value = "Disposals"
    bets_sheet.cell(3+y, 4).value = "Odds"
    bets_sheet.cell(3+y, 5).value = "Difference"
    bets_sheet.cell(3+y, 6).value = "Win?"



    singles = [["Player", "Bet", 0.0, 0.0], ["Player2", "Bet", 0.0, 0.0], ["Player3", "Bet",  0.0, 0.0], ["Player4", "Bet", 0.0, 0.0], ["Player5", "Bet", 0.0, 0.0]]
    numpy_singles = np.array(singles)
    odds_singles = [0, 0, 0, 0, 0]
    player_singles = [0, 0, 0, 0, 0]
    bets_15 = 0
    odds_15 = 0
    bets_20 = 0
    bets_25 = 0
    bets_30 = 0
    bets_35 = 0
    odds_20 = 0
    odds_25 = 0
    odds_30 = 0
    odds_35 = 0


    for i in range(2, m):
        if (sheet.cell(i, 1).value == home_array[p] or sheet.cell(i, 1).value == away_array[p]):
            if sheet.cell(i, games-1+8).value is None or sheet.cell(i, games-1+7).value is None or sheet.cell(i, games-1+6).value <60:
                fifteen = 0
                bets_15 = 0
                odds_15 = 0
            else:
                fifteen = sheet.cell(i, games-1+8).value - sheet.cell(i, games-1+7).value
                temp = sheet.cell(1, games-1+8).value.split(" -")
                bets_15 = temp[0]
                odds_15 = sheet.cell(i, games-1+8).value


            if sheet.cell(i, games-1+12).value is None or sheet.cell(i, games-1+11).value is None or sheet.cell(i, games-1+10).value <60:
                twenty = 0
                bets_20 = 0
                odds_20 = 0
            else:
                twenty = sheet.cell(i, games-1+12).value - sheet.cell(i, games-1+11).value
                temp = sheet.cell(1, games-1+12).value.split(" -")
                bets_20 = temp[0]
                odds_20 = sheet.cell(i, games-1+12).value

            if sheet.cell(i, games-1+16).value is None or sheet.cell(i, games-1+15).value is None or sheet.cell(i, games-1+14).value <60:
                twenty_five = 0
                bets_25 = 0
                odds_25 = 0
            else:
                twenty_five = sheet.cell(i, games-1+16).value - sheet.cell(i, games-1+15).value
                temp = sheet.cell(1, games-1+16).value.split(" -")
                bets_25 = temp[0]
                odds_25 = sheet.cell(i, games-1+16).value

            if sheet.cell(i, games-1+20).value is None or sheet.cell(i, games-1+19).value is None or sheet.cell(i, games-1+18).value <60:
                thirty = 0
                bets_30 = 0
                odds_30 = 0
            else:
                thirty = sheet.cell(i, games-1+20).value - sheet.cell(i, games-1+19).value
                temp = sheet.cell(1, games-1+20).value.split(" -")
                bets_30 = temp[0]
                odds_30 = sheet.cell(i, games-1+20).value

            if sheet.cell(i, games-1+24).value is None or sheet.cell(i, games-1+23).value is None or sheet.cell(i, games-1+22).value <60:
                thirty_five = 0
                bets_35 = 0
                odds_35 = 0
            else:
                thirty_five = sheet.cell(i, games-1+24).value - sheet.cell(i, games-1+23).value
                temp = sheet.cell(1, games-1+24).value.split(" -")
                bets_35 = temp[0]
                odds_35 = sheet.cell(i, games-1+24).value



            
            differences = [fifteen, twenty, twenty_five, thirty, thirty_five]
            bet = [bets_15, bets_20, bets_25, bets_30, bets_35]
            odds = [odds_15, odds_20, odds_25, odds_30, odds_35]
            
            for j in range(0, len(differences)):
                differences[j] = float(differences[j])
                temp = float(numpy_singles[4][3])
                if differences[j] > temp:
                    numpy_singles = np.delete(numpy_singles, 4, 0)
                    numpy_singles = np.vstack([numpy_singles, [sheet.cell(i, 2).value , bet[j], odds[j], differences[j]]])
                    numpy_singles = numpy_singles[numpy_singles[:, 3].argsort()][::-1]




    for k in range(0, 5):
        bets_sheet.cell(k+4+y, 2).value = numpy_singles[k][0]
        bets_sheet.cell(k+4+y, 3).value = numpy_singles[k][1]
        bets_sheet.cell(k+4+y, 4).value = numpy_singles[k][2]
        bets_sheet.cell(k+4+y, 5).value = numpy_singles[k][3]    


    #Do the multi part of it all now
    bets_sheet.cell(k+5+y, 2).value = "Multi"
    bets_sheet.cell(k+5+y, 2).font = Font(bold= True)
    bets_sheet.cell(k+5+y, 2).fill = PatternFill("solid", fgColor = "BDD7EE")
    bets_sheet.merge_cells('B' + str(k+5+y) + ':F' + str(k+5+y))
    bets_sheet.cell(k+5+y, 2).alignment = Alignment(horizontal = "center")


    multi = [["Player", "Bet", 0.0, 0.0], ["Player2", "Bet", 0.0, 0.0], ["Player3", "Bet",  0.0, 0.0], ["Player4", "Bet", 0.0, 0.0], ["Player5", "Bet", 0.0, 0.0]]
    numpy_multi = np.array(multi)
    odds_multi = [0, 0, 0, 0, 0]
    player_multi = [0, 0, 0, 0, 0]


    for i in range(2, m):
        if (sheet.cell(i, 1).value == home_array[p] or sheet.cell(i, 1).value == away_array[p]):
            if sheet.cell(i, games-1+8).value is None or sheet.cell(i, games-1+7).value is None or sheet.cell(i, games-1+6).value <90:
                fifteen = 0
                bets_15 = 0
                odds_15 = 0
            else:
                fifteen = sheet.cell(i, games-1+8).value - sheet.cell(i, games-1+7).value
                temp = sheet.cell(1, games-1+8).value.split(" -")
                bets_15 = temp[0]
                odds_15 = sheet.cell(i, games-1+8).value


            if sheet.cell(i, games-1+12).value is None or sheet.cell(i, games-1+11).value is None or sheet.cell(i, games-1+10).value <90:
                twenty = 0
                bets_20 = 0
                odds_20 = 0
            else:
                twenty = sheet.cell(i, games-1+12).value - sheet.cell(i, games-1+11).value
                temp = sheet.cell(1, games-1+12).value.split(" -")
                bets_20 = temp[0]
                odds_20 = sheet.cell(i, games-1+12).value

            if sheet.cell(i, games-1+16).value is None or sheet.cell(i, games-1+15).value is None or sheet.cell(i, games-1+14).value <90:
                twenty_five = 0
                bets_25 = 0
                odds_25 = 0
            else:
                twenty_five = sheet.cell(i, games-1+16).value - sheet.cell(i, games-1+15).value
                temp = sheet.cell(1, games-1+16).value.split(" -")
                bets_25 = temp[0]
                odds_25 = sheet.cell(i, games-1+16).value

            if sheet.cell(i, games-1+20).value is None or sheet.cell(i, games-1+19).value is None or sheet.cell(i, games-1+18).value <90:
                thirty = 0
                bets_30 = 0
                odds_30 = 0
            else:
                thirty = sheet.cell(i, games-1+20).value - sheet.cell(i, games-1+19).value
                temp = sheet.cell(1, games-1+20).value.split(" -")
                bets_30 = temp[0]
                odds_30 = sheet.cell(i, games-1+20).value

            if sheet.cell(i, games-1+24).value is None or sheet.cell(i, games-1+23).value is None or sheet.cell(i, games-1+22).value <90:
                thirty_five = 0
                bets_35 = 0
                odds_35 = 0
            else:
                thirty_five = sheet.cell(i, games-1+24).value - sheet.cell(i, games-1+23).value
                temp = sheet.cell(1, games-1+24).value.split(" -")
                bets_35 = temp[0]
                odds_35 = sheet.cell(i, games-1+24).value



            
            differences = [fifteen, twenty, twenty_five, thirty, thirty_five]
            bet = [bets_15, bets_20, bets_25, bets_30, bets_35]
            odds = [odds_15, odds_20, odds_25, odds_30, odds_35]
            
            for j in range(0, len(differences)):
                differences[j] = float(differences[j])
                temp = float(numpy_multi[4][3])
                if differences[j] > temp:
                    numpy_multi = np.delete(numpy_multi, 4, 0)
                    numpy_multi = np.vstack([numpy_multi, [sheet.cell(i, 2).value , bet[j], odds[j], differences[j]]])
                    numpy_multi = numpy_multi[numpy_multi[:, 3].argsort()][::-1]




    for k in range(0, 5):
        bets_sheet.cell(k+4+y+6, 2).value = numpy_multi[k][0]
        bets_sheet.cell(k+4+y+6, 3).value = numpy_multi[k][1]
        bets_sheet.cell(k+4+y+6, 4).value = numpy_multi[k][2]
        bets_sheet.cell(k+4+y+6, 5).value = numpy_multi[k][3]    



bets.save("Bets.xlsx")
workbook.save("Disposals Tracking.xlsx")


