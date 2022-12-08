""" TO DO

Comments

Fix up what the numbers are for each stat

Create new file that pulls out the odds
"""


import openpyxl as xl
import requests
import re
import string
import statistics as st
from scipy.stats import norm
from bs4 import BeautifulSoup

site = "https://www.basketball-reference.com/players/"
rebounds_stat = []
assists_stat = []
points_stat = []
site_array = []
row = 2

workbook = xl.load_workbook("2022 Season Data.xlsx")
points_sheet = workbook.worksheets[0]
assists_sheet = workbook.worksheets[1]
rebounds_sheet = workbook.worksheets[2]

for i in range(0, 26):
    if i !=23:
        site_array.append(site+chr(i+97)+"/")

for j in range(0, len(site_array)):
    for m in range(0, len(site_array[j])-1):
        print(site_array[m])
        webpage_response = requests.get(site_array[m])
        webpage = webpage_response.content
        soup = BeautifulSoup(webpage, "html.parser")
        bold_links = soup.find_all('strong')
        for k in range(0, len(bold_links)):
            if bold_links[k].find('a') != None:
                temp = bold_links[k].find('a').get('href').split('.html')
                site = "https://www.basketball-reference.com" + temp[0] + "/gamelog/2022"
                webpage_response2 = requests.get(site)
                webpage2 = webpage_response2.content
                soup2 = BeautifulSoup(webpage2, "html.parser")
                rebounds = soup2.find_all(attrs={'data-stat':'trb'})
                points = soup2.find_all(attrs={'data-stat':'pts'})
                assist = soup2.find_all(attrs={'data-stat':'ast'})
                name = bold_links[k].find('a').text
                for h in range(0, len(rebounds)):
                    if rebounds[h].text != "TRB":
                        rebounds_stat.append(int(rebounds[h].text))

                for h in range(0, len(rebounds)):
                    if assist[h].text != "AST":
                        assists_stat.append(int(assist[h].text))

                for h in range(0, len(rebounds)):
                    if points[h].text != "PTS":
                        points_stat.append(int(points[h].text))

                if len(points_stat)>1:
                    print(name)
                    points_average = (sum(points_stat)/len(points_stat))
                    points_std = (st.stdev(points_stat))
                    
                    points_15_percentage = (1-norm.cdf(14, points_average, points_std))*100
                    points_15_req_odds = 1/(points_15_percentage/100)

                    points_20_percentage = (1-norm.cdf(19, points_average, points_std))*100
                    points_20_req_odds = 1/(points_20_percentage/100)

                    points_25_percentage = (1-norm.cdf(24, points_average, points_std))*100
                    points_25_req_odds = 1/(points_25_percentage/100)
                    
                    points_sheet.cell(row, 1).value = name
                    
                    points_sheet.cell(row, 2).value = points_average
                    points_sheet.cell(row, 3).value = points_std
                    
                    points_sheet.cell(row, 4).value = points_15_percentage
                    points_sheet.cell(row, 5).value = points_15_req_odds
                    
                    points_sheet.cell(row, 6).value = points_20_percentage
                    points_sheet.cell(row, 7).value = points_20_req_odds
                    
                    points_sheet.cell(row, 8).value = points_25_percentage
                    points_sheet.cell(row, 9).value = points_25_req_odds
                    
                if len(assists_stat)>1:
                    assists_average = (sum(assists_stat)/len(assists_stat))
                    assists_std = (st.stdev(assists_stat))
                    
                    assists_15_percentage = (1-norm.cdf(4, assists_average, assists_std))*100
                    assists_15_req_odds = 1/(assists_15_percentage/100)

                    assists_20_percentage = (1-norm.cdf(6, assists_average, assists_std))*100
                    assists_20_req_odds = 1/(assists_20_percentage/100)
                    
                    assists_25_percentage = (1-norm.cdf(8, assists_average, assists_std))*100
                    assists_25_req_odds = 1/(assists_25_percentage/100)
                    
                    assists_sheet.cell(row, 1).value = name
                    
                    assists_sheet.cell(row, 2).value = assists_average
                    assists_sheet.cell(row, 3).value = assists_std
                    
                    assists_sheet.cell(row, 4).value = assists_15_percentage
                    assists_sheet.cell(row, 5).value = assists_15_req_odds
                    
                    assists_sheet.cell(row, 6).value = assists_20_percentage
                    assists_sheet.cell(row, 7).value = assists_20_req_odds
                    
                    assists_sheet.cell(row, 8).value = assists_25_percentage
                    assists_sheet.cell(row, 9).value = assists_25_req_odds

                    

                if len(rebounds_stat)>1:
                    rebounds_average = (sum(rebounds_stat)/len(rebounds_stat))
                    rebounds_std = (st.stdev(rebounds_stat))
                    
                    rebounds_15_percentage = (1-norm.cdf(4, rebounds_average, rebounds_std))*100
                    rebounds_15_req_odds = 1/(rebounds_15_percentage/100)

                    rebounds_20_percentage = (1-norm.cdf(6, rebounds_average, rebounds_std))*100
                    rebounds_20_req_odds = 1/(rebounds_20_percentage/100)

                    rebounds_25_percentage = (1-norm.cdf(8, rebounds_average, rebounds_std))*100
                    rebounds_25_req_odds = 1/(rebounds_25_percentage/100)
                    
                    rebounds_sheet.cell(row, 1).value = name
                    
                    rebounds_sheet.cell(row, 2).value = rebounds_average
                    rebounds_sheet.cell(row, 3).value = rebounds_std
                    
                    rebounds_sheet.cell(row, 4).value = rebounds_15_percentage
                    rebounds_sheet.cell(row, 5).value = rebounds_15_req_odds
                    
                    rebounds_sheet.cell(row, 6).value = rebounds_20_percentage
                    rebounds_sheet.cell(row, 7).value = rebounds_20_req_odds
                    
                    rebounds_sheet.cell(row, 8).value = rebounds_25_percentage
                    rebounds_sheet.cell(row, 9).value = rebounds_25_req_odds

                    row = row+1


                rebounds_stat = []
                assists_stat = []
                points_stat = []

        workbook.save("2022 Season Data.xlsx")

