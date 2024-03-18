import openpyxl as xl
import itertools
import time
import math
import numpy as np

start_time = time.time()
#Not needed if not webscraping
"""
#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)
"""

#Export CSV instead
"""
driver.get('https://play.draftstars.com.au/contest/Ci6bVJmK')

messy_players = driver.find_elements(By.XPATH, '//*[@class="css-1tnjjh5 e1rhbet73"]')
messy_salary = driver.find_elements(By.XPATH, '//*[@class="css-12bqw24"]')
players = []
salary = []

for a in messy_players:
    players.append(a.text)

for a in messy_salary:
    salary.append(a.text)

print(len(players))
print(players)
print(salary)

"""

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\NBA\\Data.xlsx")
pg = data.worksheets[1]
sg = data.worksheets[2]
sf = data.worksheets[3]
pf = data.worksheets[4]
c = data.worksheets[5]
datasheet = data.worksheets[0]

pg_array = []
sg_array = []
sf_array = []
pf_array = []
c_array = []

"""
high_data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\HIGH_Data.xlsx")
high_qb = high_data.worksheets[1]
high_rb = high_data.worksheets[2]
high_wr = high_data.worksheets[3]
high_te = high_data.worksheets[4]
high_dst = high_data.worksheets[5]
high_flex = high_data.worksheets[6]
high_datasheet = high_data.worksheets[0]

high_qb_array = []
high_rb_array = []
high_wr_array = []
high_te_array = []
high_dst_array = []
high_flex_array = []
"""

pg_array_salary = []
sg_array_salary = []
sf_array_salary = []
pf_array_salary = []
c_array_salary = []

#high_sheets = [high_qb, high_rb, high_wr, high_te, high_dst, high_flex]
sheets = [pg, sg, sf, pf, c]

arrays = [[pg_array, sg_array, sf_array, pf_array, c_array], [pg_array_salary, sg_array_salary, sf_array_salary, pf_array_salary, c_array_salary]]
#high_array = [high_qb_array, high_rb_array, high_wr_array, high_te_array, high_dst_array, high_flex_array]

max_points = 0
#high_max_points = 0

max_salary = 0
#high_max_salary = 0

team = []
#high_team = []

for h in range(0, 5):
    k = 2
    while (sheets[h].cell(k, 1).value is None) == False:
        arrays[0][h].append(sheets[h].cell(k, 3).value)
        arrays[1][h].append(sheets[h].cell(k, 6).value)
        k = k+1

actual_array = [[arrays[0][0], arrays[0][0], arrays[0][1], arrays[0][1], arrays[0][2], arrays[0][2], arrays[0][3], arrays[0][3], arrays[0][4]], [arrays[1][0], arrays[1][0], arrays[1][1], arrays[1][1], arrays[1][2], arrays[1][2], arrays[1][3], arrays[1][3], arrays[1][4]]]

"""
permutations = list(itertools.product(*actual_array[0]))

for i in range(0, len(permutations)):
    if len(permutations[i]) == len(set(permutations[i])):
        print(permutations[i])
"""
blank = [[1, 1, ["Rush"]], [2, 2, ["Not Rush"]], [3, 3, ["Quick"]], [4, 4, ["Fast"]], [5, 5, ["Slow"]], [6, 6, ["Rush"]], [7, 7, ["Not Rush"]], [8, 8, ["Quick"]], [9, 9, ["Fast"]], [10, 10, ["Slow"]]]
top_10 = np.array(blank)
count = 0
#high_blank = [[1, 1, ["Rush"]], [2, 2, ["Not Rush"]], [3, 3, ["Quick"]], [4, 4, ["Fast"]], [5, 5, ["Slow"]], [6, 6, ["Rush"]], [7, 7, ["Not Rush"]], [8, 8, ["Quick"]], [9, 9, ["Fast"]], [10, 10, ["Slow"]]]
#high_top_10 = np.array(high_blank)

for combo in itertools.product(*actual_array[0]):
    count = count + 1
    if math.floor(count/1000000) == count/1000000:
        print(count)
        end_time = time.time()
        print(end_time-start_time)

    if math.floor(count/10000000) == count/10000000:
        print(top_10)

    if len(combo) == len(set(combo)):
        print("in it")
        salary = 0
        points = 0
        #high_points = 0
        for j in range(0, 9):
            k = 2
            while combo[j] != datasheet.cell(k, 3).value:
                k = k+1
            salary = salary + datasheet.cell(k, 6).value
            points = points + datasheet.cell(k, 8).value
            #high_points = high_points + high_datasheet.cell(k, 8).value
        check = 0
        for r in range(0, 10):
            if round(points, 1) == top_10[r][1]:
                check = 1
        check_2 = 0
        for r in range(0, 10):
            if round(high_points, 1) == high_top_10[r][1]:
                check_2 = 1        
        if check == 0:
            if points > top_10[9][1] and salary<100000:
                top_10 = np.vstack([top_10, [salary, round(points, 1), combo]])
                top_10 = top_10[top_10[:, 1].argsort()][::-1]
                top_10 = np.delete(top_10, 10, 0)
        """       
        if check_2 == 0:
            if high_points > high_top_10[9][1] and salary<100000:
                high_top_10 = np.vstack([high_top_10, [salary, round(high_points, 1), combo]])
                high_top_10 = high_top_10[high_top_10[:, 1].argsort()][::-1]
                high_top_10 = np.delete(high_top_10, 10, 0)
        """

print(top_10)
print()
#print(high_top_10)

end_time = time.time()

print(end_time-start_time)





