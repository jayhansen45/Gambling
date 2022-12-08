import openpyxl as xl
import itertools
import random
import numpy as np
import time
import warnings
import math
warnings.filterwarnings("ignore")
start = time.time()

#Not needed if not webscraping
"""
#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)
"""

#Export CSV instead
"""
driver.get('https://play.draftstars.com.au/contest/Ci6bVJmK')

messy_players = driver.find_elements(By.XPATH, '//*[@class="css-1tnjjh5 e1rhbet73"]')
messy_salary = driver.find_elements(By.XPATH, '//*[@class="css-12bqw24"]')
players = []
salary = []

for a in messy_players:
    players.append(a.text)

for a in messy_salary:
    salary.append(a.text)

print(len(players))
print(players)
print(salary)

"""

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\Data.xlsx")
qb = data.worksheets[1]
rb = data.worksheets[2]
wr = data.worksheets[3]
te = data.worksheets[4]
dst = data.worksheets[5]
flex = data.worksheets[6]
datasheet = data.worksheets[0]

qb_array = []
rb_array = []
wr_array = []
te_array = []
dst_array = []
flex_array = []

high_data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\HIGH_Data.xlsx")
high_qb = high_data.worksheets[1]
high_rb = high_data.worksheets[2]
high_wr = high_data.worksheets[3]
high_te = high_data.worksheets[4]
high_dst = high_data.worksheets[5]
high_flex = high_data.worksheets[6]
high_datasheet = high_data.worksheets[0]

high_qb_array = []
high_rb_array = []
high_wr_array = []
high_te_array = []
high_dst_array = []
high_flex_array = []

qb_array_salary = []
rb_array_salary = []
wr_array_salary = []
te_array_salary = []
dst_array_salary = []
flex_array_salary = []

high_sheets = [high_qb, high_rb, high_wr, high_te, high_dst, high_flex]
sheets = [qb, rb, wr, te, dst, flex]

arrays = [[qb_array, rb_array, wr_array, te_array, dst_array, flex_array], [qb_array_salary, rb_array_salary, wr_array_salary, te_array_salary, dst_array_salary, flex_array_salary]]
high_array = [high_qb_array, high_rb_array, high_wr_array, high_te_array, high_dst_array, high_flex_array]

max_points = 0
high_max_points = 0

max_salary = 0
high_max_salary = 0

team = []
high_team = []

for h in range(0, 6):
    k = 2
    while (sheets[h].cell(k, 1).value is None) == False:
        arrays[0][h].append(sheets[h].cell(k, 3).value)
        arrays[1][h].append(sheets[h].cell(k, 6).value)
        k = k+1

actual_array = [[arrays[0][0], arrays[0][1], arrays[0][1], arrays[0][2], arrays[0][2], arrays[0][2], arrays[0][3], arrays[0][4], arrays[0][5]], [arrays[1][0], arrays[1][1], arrays[1][1], arrays[1][2], arrays[1][2], arrays[1][2], arrays[1][3], arrays[1][4], arrays[1][5]]]

"""
permutations = list(itertools.product(*actual_array[0]))

for i in range(0, len(permutations)):
    if len(permutations[i]) == len(set(permutations[i])):
        print(permutations[i])
"""
blank = []
high_blank = []


for i in range(0, 100):
    blank.append([i, i, ["blah"]])
    high_blank.append([i, i, ["blah"]])

top_100 = np.array(blank)
high_top_100 = np.array(high_blank)


combinations = []

for i in range(0, 3000000):
    if math.floor(i/100000) == i/100000:
        print(i)
        print()
        print(top_100)
        print()
        print(high_top_100)
        print()

    combo = [arrays[0][0][0], arrays[0][1][0], arrays[0][1][1], arrays[0][2][0], arrays[0][2][1], arrays[0][2][2], arrays[0][3][0], arrays[0][4][0], arrays[0][5][0]]
    if len(combo) == len(set(combo)):
        salary = 0
        points = 0
        high_points = 0
        for j in range(0, 9):
            k = 2
            while combo[j] != datasheet.cell(k, 3).value:
                k = k+1
            salary = salary + datasheet.cell(k, 6).value
            points = points + datasheet.cell(k, 8).value
            high_points = high_points + high_datasheet.cell(k, 8).value
        check = 0
        for r in range(0, 100):
            if round(points, 1) == top_100[r][1]:
                check = 1
        check_2 = 0
        for r in range(0, 100):
            if round(high_points, 1) == high_top_100[r][1]:
                check_2 = 1        
        if check == 0:
            if points > top_100[99][1] and salary<100000:
                top_100 = np.vstack([top_100, [salary, round(points, 1), combo]])
                top_100 = top_100[top_100[:, 1].argsort()][::-1]
                top_100 = np.delete(top_100, 100, 0)
                
        if check_2 == 0:
            if high_points > high_top_100[99][1] and salary<100000:
                high_top_100 = np.vstack([high_top_100, [salary, round(high_points, 1), combo]])
                high_top_100 = high_top_100[high_top_100[:, 1].argsort()][::-1]
                high_top_100 = np.delete(high_top_100, 100, 0)        

    random.shuffle(arrays[0][0])
    random.shuffle(arrays[0][1])
    random.shuffle(arrays[0][2])
    random.shuffle(arrays[0][3])
    random.shuffle(arrays[0][4])
    random.shuffle(arrays[0][5])

qb = []

for i in range(0, len(top_100)):
    qb.append(top_100[i][2][0])


qb = np.array(qb)

b, c = np.unique(qb, return_counts = True)

out = b[np.argsort(-c)]
sorte = c[np.argsort(-c)]

print(top_100)
print()
print(high_top_100)
        

end = time.time()
print(end-start)





