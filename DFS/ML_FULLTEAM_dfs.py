import openpyxl as xl
import itertools
import random
import time
import numpy as np
import warnings
warnings.filterwarnings("ignore")

start = time.time()
#Not needed if not webscraping
"""
#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)
"""

#Export CSV instead
"""
driver.get('https://play.draftstars.com.au/contest/Ci6bVJmK')

messy_players = driver.find_elements(By.XPATH, '//*[@class="css-1tnjjh5 e1rhbet73"]')
messy_salary = driver.find_elements(By.XPATH, '//*[@class="css-12bqw24"]')
players = []
salary = []

for a in messy_players:
    players.append(a.text)

for a in messy_salary:
    salary.append(a.text)

print(len(players))
print(players)
print(salary)

"""

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\Data.xlsx")
qb = data.worksheets[1]
rb = data.worksheets[2]
wr = data.worksheets[3]
te = data.worksheets[4]
dst = data.worksheets[5]
flex = data.worksheets[6]
datasheet = data.worksheets[0]

qb_array = []
rb_array = []
wr_array = []
te_array = []
dst_array = []
flex_array = []

high_data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\HIGH_Data.xlsx")
high_qb = high_data.worksheets[1]
high_rb = high_data.worksheets[2]
high_wr = high_data.worksheets[3]
high_te = high_data.worksheets[4]
high_dst = high_data.worksheets[5]
high_flex = high_data.worksheets[6]
high_datasheet = high_data.worksheets[0]

high_qb_array = []
high_rb_array = []
high_wr_array = []
high_te_array = []
high_dst_array = []
high_flex_array = []

qb_array_salary = []
rb_array_salary = []
wr_array_salary = []
te_array_salary = []
dst_array_salary = []
flex_array_salary = []

high_sheets = [high_qb, high_rb, high_wr, high_te, high_dst, high_flex]
sheets = [qb, rb, wr, te, dst, flex]

arrays = [qb_array, rb_array, wr_array, te_array, dst_array, flex_array]
high_array = [high_qb_array, high_rb_array, high_wr_array, high_te_array, high_dst_array, high_flex_array]

max_points = 0
high_max_points = 0

max_salary = 0
high_max_salary = 0

#team = [["", 1], ["", 2], ["", 3], ["", 4], ["", 5], ["", 6], ["", 7], ["", 8], ["", 9]]
team = []
high_team = []

for h in range(0, 6):
    k = 2
    while (sheets[h].cell(k, 1).value is None) == False:
        arrays[h].append(sheets[h].cell(k, 3).value)
        k = k+1

actual_array = [[arrays[0], arrays[1], arrays[1], arrays[2], arrays[2], arrays[2], arrays[3], arrays[4], arrays[5]], [arrays[0], arrays[1], arrays[1], arrays[2], arrays[2], arrays[2], arrays[3], arrays[4], arrays[5]]]

"""
permutations = list(itertools.product(*actual_array[0]))

for i in range(0, len(permutations)):
    if len(permutations[i]) == len(set(permutations[i])):
        print(permutations[i])
"""
blank = []
high_blank = []
index_array = []
total_points = 0

for i in range(0, 50):
    blank.append([i, i, ["blah", "f", "d"]])
    high_blank.append([i, i, ["blah", "d", "r"]])

top_100 = np.array(blank)
high_top_100 = np.array(high_blank)

spent = 0
counts = []
combinations = []

for w in range(0, len(actual_array[0])):
    top_100 = np.array(blank)
    high_top_100 = np.array(high_blank)
    for i in range(0, 10000):
        combo = [arrays[0][0], arrays[1][0], arrays[1][1], arrays[2][0], arrays[2][1], arrays[2][2], arrays[3][0], arrays[4][0], arrays[5][0]]
        for t in range(0, len(index_array)):
            combo = np.delete(combo, index_array[t])

        tops = []
        counts = []
        if len(combo) == len(set(combo)):
            salary = 0
            points = 0
            high_points = 0
            for j in range(0, len(combo)):
                k = 2
                while combo[j] != datasheet.cell(k, 3).value:
                    k = k+1
                salary = salary + datasheet.cell(k, 6).value
                points = points + datasheet.cell(k, 8).value
                high_points = high_points + high_datasheet.cell(k, 8).value
            check = 0
            for r in range(0, 50):
                if round(points, 1) == top_100[r][1]:
                    check = 1
            check_2 = 0
            for r in range(0, 50):
                if round(high_points, 1) == high_top_100[r][1]:
                    check_2 = 1        
            if check == 0:
                if points > top_100[49][1] and salary<100000-spent:
                    top_100 = np.vstack([top_100, [salary, round(points, 1), combo]])
                    top_100 = top_100[top_100[:, 1].argsort()][::-1]
                    top_100 = np.delete(top_100, 50, 0)
                    
            if check_2 == 0:
                if high_points > high_top_100[49][1] and salary<100000-spent:
                    high_top_100 = np.vstack([high_top_100, [salary, round(high_points, 1), combo]])
                    high_top_100 = high_top_100[high_top_100[:, 1].argsort()][::-1]
                    high_top_100 = np.delete(high_top_100, 50, 0)        

        random.shuffle(arrays[0])
        random.shuffle(arrays[1])
        random.shuffle(arrays[2])
        random.shuffle(arrays[3])
        random.shuffle(arrays[4])
        random.shuffle(arrays[5])

    for p in range(0, len(combo)):
        temp = []
        for i in range(0, len(top_100)):
            check3 = 0
            for b in range(0, len(team)):
                if top_100[i][2][p] == team[b][0]:
                    check3 = 1
            if check3 == 0:
                temp.append(top_100[i][2][p])


        player = np.array(temp)
        b, c = np.unique(player, return_counts = True)
        out = b[np.argsort(-c)]
        sorte = c[np.argsort(-c)]

        tops.append(out[0])
        counts.append(sorte[0])
    counts = np.array(counts)
    index = np.argmax(counts)

    team.append([tops[index], counts[index], index])
    #team[index] = [tops[index], counts[index]]
    player_salary = 0
    player_points = 0
    k = 2
    while tops[index] != datasheet.cell(k, 3).value:
        k = k + 1
    player_salary = datasheet.cell(k, 6).value
    points = points + datasheet.cell(k, 8).value
    spent = spent + player_salary
    total_points = total_points + player_points
    index_array.append(index)
    print(team)
    
print(team)
print(spent, total_points)
end = time.time()

print(end-start)




