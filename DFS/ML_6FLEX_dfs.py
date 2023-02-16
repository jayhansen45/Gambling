import openpyxl as xl
import itertools
import time
import math
import random
import numpy as np
import collections
import warnings
warnings.filterwarnings("ignore")

start_time = time.time()

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\Data.xlsx")
flex = data.worksheets[6]

high_data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\HIGH_Data.xlsx")
high_flex = high_data.worksheets[6]

blank = []
high_blank = []
                                                                                               
for i in range(0, 50):
    blank.append([i, i, ["blah"]])
    high_blank.append([i, i, ["blah"]])            

top_50 = np.array(blank)
high_top_50 = np.array(high_blank)

combo = []
combo_salary = []
randoms = []
projections = []
spent = 0
check = 0
player = np.empty

    
#Loops through the stuff 6 times so that we get all 6 flex positions
for w in range(0, 1):
    for j in range(0, 10000):
        combo = []
        combo_salary = []
        projections = []
        randoms = []
        check= 0
        
        #Creates an array of random names
        while len(randoms)<6:
            index = random.randint(2, 21)
            if index not in randoms:
                randoms.append(index)
                combo.append(flex.cell(index, 3).value)
                combo_salary.append(flex.cell(index, 6).value)
                projections.append(flex.cell(index, 8).value)

        #Appends to the top 50 list if it is less than 80,000
        for r in range(0, 50):
            if top_50[r][1] == round(sum(projections), 1):
                check = 1
                
        if sum(combo_salary)<80000-spent and check == 0:
            top_50 = np.vstack([top_50, [sum(combo_salary), round(sum(projections), 1), combo]])
            top_50 = top_50[top_50[:, 1].argsort()][::-1]
            top_50 = np.delete(top_50, 50, 0)

    for i in range(0, len(top_50)):
        player = np.append(player, top_50[i][2])

    uniq = ["name", 1]
    unique = np.array(uniq)
    
    for h in range(2, 22):
        temp = flex.cell(h, 3).value
        count = 0
        for p in player:
            if temp == p:
                count = count + 1
        unique = np.vstack([unique, [temp, int(count)]])
    print(type(unique[0][1]))
    print(unique[0][1])
    print(unique)
    unique = unique[unique[:, 1].argsort()][::-1]
    print(unique)
    

end_time=time.time()
print(end_time-start_time)






