import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import shutil
import os
from datetime import datetime, timedelta, date
import time
from selenium.webdriver.common.action_chains import ActionChains
import itertools


#Not needed if not webscraping
"""
#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)
"""

#Export CSV instead
"""
driver.get('https://play.draftstars.com.au/contest/Ci6bVJmK')

messy_players = driver.find_elements(By.XPATH, '//*[@class="css-1tnjjh5 e1rhbet73"]')
messy_salary = driver.find_elements(By.XPATH, '//*[@class="css-12bqw24"]')
players = []
salary = []

for a in messy_players:
    players.append(a.text)

for a in messy_salary:
    salary.append(a.text)

print(len(players))
print(players)
print(salary)

"""

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\Data.xlsx")
qb = data.worksheets[1]
rb = data.worksheets[2]
wr = data.worksheets[3]
te = data.worksheets[4]
dst = data.worksheets[5]
flex = data.worksheets[6]
datasheet = data.worksheets[0]

qb_array = []
rb_array = []
wr_array = []
te_array = []
dst_array = []
flex_array = []

qb_array_salary = []
rb_array_salary = []
wr_array_salary = []
te_array_salary = []
dst_array_salary = []
flex_array_salary = []

sheets = [qb, rb, wr, te, dst, flex]
arrays = [[qb_array, rb_array, wr_array, te_array, dst_array, flex_array], [qb_array_salary, rb_array_salary, wr_array_salary, te_array_salary, dst_array_salary, flex_array_salary]]

max_points = 0
max_salary = 0
team = []

for h in range(0, 6):
    k = 2
    while (sheets[h].cell(k, 1).value is None) == False:
        arrays[0][h].append(sheets[h].cell(k, 3).value)
        arrays[1][h].append(sheets[h].cell(k, 6).value)
        k = k+1

actual_array = [[arrays[0][0], arrays[0][1], arrays[0][1], arrays[0][2], arrays[0][2], arrays[0][2], arrays[0][3], arrays[0][4], arrays[0][5]], [arrays[1][0], arrays[1][1], arrays[1][1], arrays[1][2], arrays[1][2], arrays[1][2], arrays[1][3], arrays[1][4], arrays[1][5]]]
"""
permutations = list(itertools.product(*actual_array[0]))

for i in range(0, len(permutations)):
    if len(permutations[i]) == len(set(permutations[i])):
        print(permutations[i])
"""

for combo in itertools.product(*actual_array[0]):
    if len(combo) == len(set(combo)):
        print(combo)
        salary = 0
        points = 0
        for j in range(0, 9):
            k = 2
            while combo[j] != datasheet.cell(k, 3).value:
                k = k+1
            salary = salary + datasheet.cell(k, 6).value
            points = points + datasheet.cell(k, 8).value
        if points > max_points and salary<100000:
            max_salary = salary
            max_points = points
            team = combo

print(max_salary, max_points, team)










