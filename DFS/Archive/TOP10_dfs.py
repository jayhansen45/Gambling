import openpyxl as xl
import itertools
import numpy as np


#Not needed if not webscraping
"""
#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='106.0.5249.61').install(), options = chrome_options)
"""

#Export CSV instead
"""
driver.get('https://play.draftstars.com.au/contest/Ci6bVJmK')

messy_players = driver.find_elements(By.XPATH, '//*[@class="css-1tnjjh5 e1rhbet73"]')
messy_salary = driver.find_elements(By.XPATH, '//*[@class="css-12bqw24"]')
players = []
salary = []

for a in messy_players:
    players.append(a.text)

for a in messy_salary:
    salary.append(a.text)

print(len(players))
print(players)
print(salary)

"""

data = xl.load_workbook("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\DFS\\Data.xlsx")
qb = data.worksheets[1]
rb = data.worksheets[2]
wr = data.worksheets[3]
te = data.worksheets[4]
dst = data.worksheets[5]
flex = data.worksheets[6]
datasheet = data.worksheets[0]

qb_array = []
rb_array = []
wr_array = []
te_array = []
dst_array = []
flex_array = []

qb_array_salary = []
rb_array_salary = []
wr_array_salary = []
te_array_salary = []
dst_array_salary = []
flex_array_salary = []

sheets = [qb, rb, wr, te, dst, flex]
arrays = [[qb_array, rb_array, wr_array, te_array, dst_array, flex_array], [qb_array_salary, rb_array_salary, wr_array_salary, te_array_salary, dst_array_salary, flex_array_salary]]

max_points = 0
max_salary = 0
team = []

for h in range(0, 6):
    k = 2
    while (sheets[h].cell(k, 1).value is None) == False:
        arrays[0][h].append(sheets[h].cell(k, 3).value)
        arrays[1][h].append(sheets[h].cell(k, 6).value)
        k = k+1

actual_array = [[arrays[0][0], arrays[0][1], arrays[0][1], arrays[0][2], arrays[0][2], arrays[0][2], arrays[0][3], arrays[0][4], arrays[0][5]], [arrays[1][0], arrays[1][1], arrays[1][1], arrays[1][2], arrays[1][2], arrays[1][2], arrays[1][3], arrays[1][4], arrays[1][5]]]
"""
permutations = list(itertools.product(*actual_array[0]))

for i in range(0, len(permutations)):
    if len(permutations[i]) == len(set(permutations[i])):
        print(permutations[i])
"""
blank = [[1, 1, ["Rush"]], [2, 2, ["Not Rush"]], [3, 3, ["Quick"]], [4, 4, ["Fast"]], [5, 5, ["Slow"]], [6, 6, ["Rush"]], [7, 7, ["Not Rush"]], [8, 8, ["Quick"]], [9, 9, ["Fast"]], [10, 10, ["Slow"]]]
top_5 = np.array(blank)

print(top_5)
for combo in itertools.product(*actual_array[0]):
    if len(combo) == len(set(combo)):
        print(combo)
        salary = 0
        points = 0
        for j in range(0, 9):
            k = 2
            while combo[j] != datasheet.cell(k, 3).value:
                k = k+1
            salary = salary + datasheet.cell(k, 6).value
            points = points + datasheet.cell(k, 8).value
        check = 0
        for r in range(0, 10):
            if round(points, 1) == top_5[r][1]:
                check = 1
        if check == 0:
            if points > top_5[9][1] and salary<100000:
                top_5 = np.vstack([top_5, [salary, round(points, 1), combo]])
                top_5 = top_5[top_5[:, 1].argsort()][::-1]
                top_5 = np.delete(top_5, 10, 0)

print(top_5)







