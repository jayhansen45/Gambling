""" TO DO
Maybe outlier if I can be fucked

Fix for when it isn't 14 games
    Move the create headings thing inside the for loop and then everything can be referenced after this instead of using "games"
    Pull out the round number from somewhere?

Try again with autosizing

Create a "STEPS" Thing at the top of this code

Fix issue about the ASCII letter thing
    if statement that checks how much t is?
    Should be a better way of doing this

    **** THIS WILL WORK ****
    Can save the values for each that needs to be remembered when the heading is written
    Use the get_column_letter thing
"""

import openpyxl as xl
import string
import requests
import statistics
import numpy as np
import warnings
from scipy.stats import norm
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter

#Ignore warnings and creates the required variables
warnings.filterwarnings("ignore")
urls = ['https://afltables.com/afl/stats/teams/adelaide/2022_gbg.html', 'https://afltables.com/afl/stats/teams/brisbanel/2022_gbg.html', 'https://afltables.com/afl/stats/teams/carlton/2022_gbg.html', 'https://afltables.com/afl/stats/teams/collingwood/2022_gbg.html', 'https://afltables.com/afl/stats/teams/essendon/2022_gbg.html', 'https://afltables.com/afl/stats/teams/fremantle/2022_gbg.html', 'https://afltables.com/afl/stats/teams/geelong/2022_gbg.html', 'https://afltables.com/afl/stats/teams/goldcoast/2022_gbg.html', 'https://afltables.com/afl/stats/teams/gws/2022_gbg.html', 'https://afltables.com/afl/stats/teams/hawthorn/2022_gbg.html', 'https://afltables.com/afl/stats/teams/melbourne/2022_gbg.html', 'https://afltables.com/afl/stats/teams/kangaroos/2022_gbg.html', 'https://afltables.com/afl/stats/teams/padelaide/2022_gbg.html', 'https://afltables.com/afl/stats/teams/richmond/2022_gbg.html', 'https://afltables.com/afl/stats/teams/stkilda/2022_gbg.html', 'https://afltables.com/afl/stats/teams/bullldogs/2022_gbg.html', 'https://afltables.com/afl/stats/teams/westcoast/2022_gbg.html', 'https://afltables.com/afl/stats/teams/swans/2022_gbg.html']
depth = 0
games = 15

workbook = xl.Workbook()
worksheet = workbook.worksheets[0]


#Creates the top row for all of the columns in the file
worksheet.cell(1, 1).value = "Team"
worksheet.column_dimensions['B'].width = 15
worksheet.cell(1, 2).value = "Players"
worksheet.column_dimensions['A'].width = 15
for i in range(3, games+3):
    worksheet.cell(1, i).value = "Round " + str(i-2)
worksheet.cell(1, i+1).value = "Average"
worksheet.cell(1, i+2).value = "StDev"
worksheet.cell(1, i+3).value = "15+ Percentage"
worksheet.cell(1, i+4).value = "15+ Req Odds"
worksheet.cell(1, i+5).value = "15+ Odds"
worksheet.cell(1, i+6).value = "15+ Difference"

worksheet.cell(1, i+7).value = "20+ Percentage"
worksheet.cell(1, i+8).value = "20+ Req Odds"
worksheet.cell(1, i+9).value = "20+ Odds"
worksheet.cell(1, i+10).value = "20+ Difference"

worksheet.cell(1, i+11).value = "25+ Percentage"
worksheet.cell(1, i+12).value = "25+ Req Odds"
worksheet.cell(1, i+13).value = "25+ Odds"
worksheet.cell(1, i+14).value = "25+ Difference"

worksheet.cell(1, i+15).value = "30+ Percentage"
worksheet.cell(1, i+16).value = "30+ Req Odds"
worksheet.cell(1, i+17).value = "30+ Odds"
worksheet.cell(1, i+18).value = "30+ Difference"

worksheet.cell(1, i+19).value = "35+ Percentage"
worksheet.cell(1, i+20).value = "35+ Req Odds"
worksheet.cell(1, i+21).value = "35+ Odds"
worksheet.cell(1, i+22).value = "35+ Difference"
worksheet.cell(1, i+23).value = "Max Difference"

j=i
i=0

#Widens all of the columns
for i in range(1, j+23+1):
    if i > 26:
        worksheet.column_dimensions['A'+ string.ascii_uppercase[i-1-26]].width = 15       
    else:
        worksheet.column_dimensions[string.ascii_uppercase[i-1]].width = 15


#Loops through all of the sites for each of the teams
for site in urls:
    print(site)
    webpage_response = requests.get(site)
    webpage = webpage_response.content
    soup = BeautifulSoup(webpage, "html.parser")

    #Finds the correct attributes
    table = soup.find_all(attrs={'class':'sortable'})
    team = soup.find('h2').text
    team = team.split(' -')


    #Pulls out the disposals and the time on ground
    messy_disposals = table[0].find_all('td')
    messy_tog = table[21].find_all('td')

    disposals = []
    tog = []


    #Gets text of the disposals and time on ground and puts in array
    for a in messy_disposals:
        if (a.text == '\xa0' or a.text == '-'):
            a = '0'
            disposals.append(a)
        else:
            disposals.append(a.text)

    for a in messy_tog:
        if (a.text == '\xa0' or a.text == '-'):
            a = '0'
            tog.append(a)
        else:
            tog.append(a.text)

    i = 1

    #Finds total number of games
    while disposals[i].isnumeric():
        i=i+1

    games = i-2
    
    gross_players = []

    i=0

    #Pulls out the players from the large array
    while (i < len(disposals)):
        gross_players.append(disposals.pop(i))
        tog.pop(i)
        i=i+games+1

    players = []

    #Puts the players in the same format as the odds site
    for i in range(0, len(gross_players)):
        gross_players[i] = gross_players[i].split(', ')
        players.append(gross_players[i][1] + " " + gross_players[i][0])

    i=games

    #Gets rid of the totals in the array
    while (i < len(disposals)):
        disposals.pop(i)
        tog.pop(i)
        i=i+games

    b=0
    c=0


    #Makes the array a numpy array
    disposals_numpy = np.array(disposals)

    #Reshapes it so that it is a 2D array
    disp = disposals_numpy.reshape(len(players), games)

    tog_numpy = np.array(tog)

    time = tog_numpy.reshape(len(players), games)    

    sum = 0


    #Pastes all of the relevant values
    for a in range(0, len(players)):
        sum = 0
        count = 0
        values = []
        worksheet.cell(a+2+depth, 1).value = team[0]
        worksheet.cell(a+2+depth, 2).value = players[a]
        #Pastes all of the disposals in the sheet with the parameters
        for t in range(0, games):
            worksheet.cell(a+2+depth, t+3).value = int(disp[a][t])
            if int(time[a][t]) < 60:
                worksheet.cell(a+2+depth, t+3).value = 0
                
            if worksheet.cell(a+2+depth, t+3).value != 0:
                sum = sum + int(disp[a][t])
                count = count + 1
                values.append(int(disp[a][t]))
        if count > 8:
            avg = sum/count
        else:
            avg = 0

        worksheet.cell(a+2+depth, t+4).value = avg
        values_numpy = np.array(values)
        if avg == 0:
            worksheet.cell(a+2+depth, t+5).value = 0
        else:
            worksheet.cell(a+2+depth, t+5).value = np.std(values_numpy)

        worksheet.cell(a+2+depth, t+6).value = (1-norm.cdf(14, avg, worksheet.cell(a+2+depth, t+5).value))*100
        worksheet.cell(a+2+depth, t+7).value = 1/(worksheet.cell(a+2+depth, t+6).value/100)
        worksheet.cell(a+2+depth, t+9).value = "=" + chr(t+8+64)+str((a+2+depth))+"-"+chr(t+7+64) + str(a+2+depth)

        
        worksheet.cell(a+2+depth, t+10).value = (1-norm.cdf(19, avg, worksheet.cell(a+2+depth, t+5).value))*100
        worksheet.cell(a+2+depth, t+11).value = 1/(worksheet.cell(a+2+depth, t+10).value/100)
        worksheet.cell(a+2+depth, t+13).value = "=" + chr(t+12+64)+str((a+2+depth))+"-"+chr(t+11+64)+str(a+2+depth)

        worksheet.cell(a+2+depth, t+14).value = (1-norm.cdf(24, avg, worksheet.cell(a+2+depth, t+5).value))*100
        worksheet.cell(a+2+depth, t+15).value = 1/(worksheet.cell(a+2+depth, t+14).value/100)
        worksheet.cell(a+2+depth, t+17).value = ("=A" + chr(t+16+64-26)+str((a+2+depth))+"-"+"A" + chr(t+15+64-26)+str(a+2+depth))

        worksheet.cell(a+2+depth, t+18).value = (1-norm.cdf(29, avg, worksheet.cell(a+2+depth, t+5).value))*100
        worksheet.cell(a+2+depth, t+19).value = 1/(worksheet.cell(a+2+depth, t+18).value/100)
        worksheet.cell(a+2+depth, t+21).value = ("=A" + chr(t+20+64-26)+str((a+2+depth))+"-"+"A" + chr(t+19+64-26)+str(a+2+depth))
        
        worksheet.cell(a+2+depth, t+22).value = (1-norm.cdf(34, avg, worksheet.cell(a+2+depth, t+5).value))*100
        worksheet.cell(a+2+depth, t+23).value = 1/(worksheet.cell(a+2+depth, t+22).value/100)
        worksheet.cell(a+2+depth, t+25).value = ("=A" + chr(t+24+64-26)+str((a+2+depth))+"-"+"A" + chr(t+23+64-26)+str(a+2+depth))

        if avg > 0:
            worksheet.cell(a+2+depth, t+26).value = ("=MAX("+ chr(t+9+64) + str(a+2+depth) + ", A" + chr(t+13+64-26) + str(a+2+depth) + ", A" + chr(t+17+64-26) + str(a+2+depth) + ", A" + chr(t+21+64-26) + str(a+2+depth) + ", A"+ chr(t+25+64-26) + str(a+2+depth) + ")")        
        else:
            worksheet.cell(a+2+depth, t+26).value = 0

    depth = depth + len(players)


        
workbook.save("Disposals Tracking.xlsx")
