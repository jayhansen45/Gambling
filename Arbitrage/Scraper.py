import requests
import openpyxl as xl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import time
import shutil
import os
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
#chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(executable_path=r"C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Other\\Chrome Driver\\chromedriver.exe", options = chrome_options)
#driver = webdriver.Chrome(ChromeDriverManager(version='118.0.5993.70').install(), options = chrome_options)

filedate=date.today()
day = (filedate.strftime('%d/%m/%Y'))
day_site = (filedate.strftime('%A'))
filedate = (filedate.strftime('%A %#d %B %Y'))

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Arbitrage.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\Historical\\Arbitrage.xlsx"
newName = "Arbitrage " + filedate
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\Arbitrage.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\" + newName + ".xlsx")

workbook = xl.load_workbook(filename)

print("Enter the number for the sports you want to check. End with Done")

#Prints the sports options
sports_text = ["1. AFL", "2. NRL", "3. NHL", "4. NBA", "5. MLB", "6. IPL", "7. EPL"]

for i in sports_text:
    print(i)

#Gets input from the user
sport_input = ""
sports = []

while sport_input != "Done":
    sport_input = input()
    sports.append(sport_input)

sports.pop()

for sport in sports:
    sheet = workbook.worksheets[int(sport)-1]
    
    #Finds the first row that hasn't been used
    m=1

    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break

    for start in range(1, 1048576):
        if sheet.cell(start, 3).value is None:
            break
        
    for end in range(1, 1048576):
        if sheet.cell(end, 1).value is None:
            break


    #defines each of the sports as functions
    def sportsbet(sports):
        print("Sportsbet " + sports_text[int(sport)-1].split(" ")[1])

        #options to access values in each site
        site_text = ["https://www.sportsbet.com.au/betting/australian-rules/afl", "https://www.sportsbet.com.au/betting/rugby-league/nrl", "https://www.sportsbet.com.au/betting/ice-hockey-us/nhl-games", "https://www.sportsbet.com.au/betting/basketball-us/nba", "https://www.sportsbet.com.au/betting/baseball/mlb-matches", "https://www.sportsbet.com.au/betting/cricket/indian-premier-league", "https://www.sportsbet.com.au/betting/soccer/united-kingdom/english-premier-league"]
        bets = ["Head to Head", "Head to Head", "Moneyline", "Match Betting", "Money Line", "Match Betting", "Win-Draw-Win"]
        driver.get(site_text[int(sport)-1])
        global sportsbet_teams
        sportsbet_teams = []
        

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[int(sport)-1] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, '//*[@class="size12_fq5j3k2 normal_fgzdi7m caption_f4zed5e"]')


        for a in messy_teams:
            sportsbet_teams.append(a.text)

        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []

        for a in messy_more_odds:
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))

        if (sport !=7):
            for count, teams_list in enumerate(sportsbet_teams):
                sheet.cell(count + m, 1).value = sportsbet_teams[count]
                sheet.cell(count + m, 2).value = more_odds[count]
        else:
            for count, teams_list in enumerate(sportsbet_teams):
                sheet.cell(count + m, 1).value = sportsbet_teams[count]
                sheet.cell(count + m, 2).value = more_odds[count]

    #Find new column for it
    def bet_deluxe(sports):
        print("Bet Deluxe  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["https://betdeluxe.com.au/sports/australian-rules/toyota-afl-premiership-1000064", "https://betdeluxe.com.au/sports/rugby-league/telstra-premiership-1000076", "https://betdeluxe.com.au/sports/ice-hockey/national-hockey-league-1000082", "https://betdeluxe.com.au/sports/basketball/nba-1000059", "https://betdeluxe.com.au/sports/baseball/major-league-baseball-1000054", "https://betdeluxe.com.au/sports/cricket/tata-indian-premier-league-1001183", "https://betdeluxe.com.au/sports/football/england-premier-league-1000009"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="css-md7rl8-Text-Text ea6hjv30"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        
        if(sport=="1"):
            temp = []
            for a in messy_odds[::2]:
                temp.append(a)
            messy_odds = temp

        temp = []
        
        if (sport == "6"):
            for game in messy_teams[0:((len(sportsbet_teams)//2)+1):1]:
                temp.append(game.text.split(" v ")[0])
                temp.append(game.text.split(" v ")[1])
        elif (sport == "7"):
            for game in messy_teams[0:((len(sportsbet_teams)//2)+1):1]:
                temp.append(game.text.split(" v ")[0])
                temp.append("Draw")
                temp.append(game.text.split(" v ")[1])
        else:
            for game in messy_teams:
                temp.append(game.text.split(" v ")[0])
                temp.append(game.text.split(" v ")[1])

        convert = {"Brisbane Lions" : "Brisbane", "Nottingham Forest" : "Nottm Forest", "Brighton & Hove Albion" : "Brighton", "Manchester City" : "Man City", "AFC Bournemouth" : "Bournemouth", "Newcastle United" : "Newcastle", "Tottenham Hotspur" : "Tottenham", "Manchester United" : "Man Utd", "Wolverhampton Wanderers" : "Wolves", "Leeds United" : "Leeds", "Leicester City" : "Leicester"}

        #This is so fucking gross   
        for count, a in enumerate(temp):
                if a in convert:
                    teams.append(convert[a])
                else:
                    teams.append(a)

                if messy_odds[count].text == '0':
                    teams.append(0)
                elif messy_odds[count].text == "Suspended":
                    teams.append(0)
                else:
                    teams.append(float(messy_odds[count].text))        

        #does the match and then saves values in the right cells
        if (sport!="7"):           
            for i in range(0, len(teams), 2):
                for j in range(0, len(sportsbet_teams)):
                    if teams[i] == sportsbet_teams[j]:
                        sheet.cell(j+m, 11).value = teams[i+1]
        #Not fucking working
        else:
            for count, i in enumerate(sportsbet_teams):
                for num, j in enumerate(teams):
                    if (i == j) and (j != "Draw") and (num != (len(teams)-2)):
                        if count%3 == 0 and sportsbet_teams[count+2] == teams[num+2]:
                            sheet.cell(count+m, 3).value = teams[num+1]
                            sheet.cell(count+m+1, 3).value = teams[num+3]
                        elif count%3 == 2 and sportsbet_teams[count-2] == teams[num-2]:
                            sheet.cell(count+m, 3).value = teams[num+1]

            
            """
            for count, i in enumerate(sportsbet_teams):
                for num, j in enumerate(teams):
                    if (i == j) and (j != "Draw") and (num != (len(teams)-2)) and (sheet.cell(count+m, 3).value) is None:
                        if count%3 == 0:
                            sheet.cell(count+m, 3).value = teams[num+1]
                            sheet.cell(count+m+1, 3).value = teams[num+3]
                        else: 
                            sheet.cell(count+m, 3).value = teams[num+1]
            """
                            

    def betr(sports):
        print("BetR  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "https://betr.com.au/sportsbook#/sport/1006/competition/1000623/1002652", "https://betr.com.au/sportsbook#/sport/13/competition/1000649/1003042"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="OddsButton_priceType__ROL+V SportsMarketCard_priceTypeText__iiWi2"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@data-testid="win_odds_button_price_value"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 3).value = teams[i+1]

    def bet365(sports):
        print("Bet365  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.bet365.com.au/#/AC/B18/C20604387/D48/E1453/F10/"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="scb-ParticipantFixtureDetailsHigherBasketball_Team"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="sac-ParticipantOddsOnly50OTB_Odds"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 7).value = teams[i+1]

    def bet_nation(sports):
        print("Bet Nation  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://betnation.com.au/sports/basketball/nba-1000003"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="css-ggkqa9-sportsStyles-styled-sportsStyles__OuterTitleContainer-sportsStyles-styled errtm9l20"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="css-1jslun9-sportsStyles-styled-sportsStyles__BetText-sportsStyles-styled errtm9l12"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 8).value = teams[i+1]
        """
    def colossalbet(sports):
        print("Colossal Bet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.colossalbet.com.au/sport/all/basketball/nba"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="mb-2 text-black font-14 font-medium"]')

        print(messy_teams)

        for count, t in enumerate(messy_teams):
            messy_teams[count] = t.text.split(' (')[0]

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="text-black font-16 font-extrabold"]')

        for count, a in enumerate(messy_teams):
            teams.append(a)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 12).value = teams[i+1]
        """

    def okebet(sports):
        print("OkeBet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://okebet.com.au/sports/basketball/nba-1000002"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []
        temp = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="errtm9l15 css-115g8u-Text-Text-sportsStyles-styled-sportsStyles__SelectionTitleText-sportsStyles-styled ea6hjv30"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for w in messy_odds[::2]:
            temp.append(w)
        
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if temp[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(temp[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 13).value = teams[i+1]

    def realbookie(sports):
        print("Real Bookie  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.realbookie.com.au/sports/basketball/nba-1000003"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []
        temp = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="errtm9l15 css-115g8u-Text-Text-sportsStyles-styled-sportsStyles__SelectionTitleText-sportsStyles-styled ea6hjv30"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for w in messy_odds[::2]:
            temp.append(w)
        
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if temp[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(temp[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 14).value = teams[i+1]

    def unibet(sports):
        print("Unibet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.unibet.com.au/betting/sports/filter/basketball/nba/all/matches"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="ca197 isExpandedView"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="_8e013"]')

        #Include if there is a featured bets thing
        #messy_odds = messy_odds[6:]


        temp = []

        for count, t in enumerate(messy_odds[3::4]):
            messy_odds.pop(count)

        for count, t in enumerate(messy_odds[2::3]):
            messy_odds.pop(count)


        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            teams.append(float(messy_odds[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 9).value = teams[i+1]


    def getsetbet(sports):
        print("GetSetBet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.getsetbet.com.au/sports/basketball/nba-1000042"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []
        temp = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="errtm9l15 css-115g8u-Text-Text-sportsStyles-styled-sportsStyles__SelectionTitleText-sportsStyles-styled ea6hjv30"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for w in messy_odds[::2]:
            temp.append(w)
        
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if temp[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(temp[count].text))


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 10).value = teams[i+1]

    def boombet(sports):
        print("BoomBet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["https://www.boombet.com.au/sport-menu/Sport/Australian%20Rules/AFL", "", "https://www.boombet.com.au/sport-menu/Sport/Ice%20Hockey/US%20NHL%20Regular%20Season-22", "https://www.boombet.com.au/sport-menu/Sport/Basketball/US%20NBA", "https://www.boombet.com.au/sport-menu/Sport/Baseball/US%20Major%20League%20Baseball%20Regular%20Season-23", "https://www.boombet.com.au/sport-menu/Sport/Cricket/India%20Premier%20League"]
        bets = ["", "", ""]
        temp = []
        odds_temp = []
        if (sport == "3"):
            driver.get("https://www.boombet.com.au/sport-menu/Sport/Ice%20Hockey/US%20NHL%20Eastern%20Conference%20Round%201-22")
            time.sleep(10)
            messy_teams1 = driver.find_elements(By.XPATH, '//*[@class="teamName d-block d-md-flex pb-1"]')
            messy_odds1 = driver.find_elements(By.XPATH, '//*[@class="oddsValue d-block d-md-flex"]')
            for a in messy_teams1:
                temp.append(a.text)

            for b in messy_odds1:
                odds_temp.append(b.text)
            
            driver.get("https://www.boombet.com.au/sport-menu/Sport/Ice%20Hockey/US%20NHL%20Western%20Conference%20Round%201-22")
            time.sleep(10)
            messy_teams2 = driver.find_elements(By.XPATH, '//*[@class="teamName d-block d-md-flex pb-1"]')
            messy_odds2 = driver.find_elements(By.XPATH, '//*[@class="oddsValue d-block d-md-flex"]')
            
            for a in messy_teams2:
                temp.append(a.text)

            for b in messy_odds2:
                odds_temp.append(b.text)
        else: 
            driver.get(site_text[int(sport)-1])
            time.sleep(10)
            messy_teams = driver.find_elements(By.XPATH, '//*[@class="teamName d-block d-md-flex pb-1"]')
            messy_odds = driver.find_elements(By.XPATH, '//*[@class="oddsValue d-block d-md-flex"]')

            for a in messy_teams:
                temp.append(a.text)

            for b in messy_odds:
                odds_temp.append(b.text)

        teams = []


        if sport !=1:
            for count, a in enumerate(temp):
                if a == "St Louis Cardinals":
                    teams.append("St. Louis Cardinals")
                else:
                    teams.append(a)
                if odds_temp[count] == "Suspended":
                    teams.append(0)
                else:
                    teams.append(float(odds_temp[count]))


        if sport == "1":    
            for count, a in enumerate(temp):
                if a == "Western Bulldogs":
                    teams.append(a)
                else:
                    teams.append(a.rsplit(" vs ", 1)[0])
                if odds_temp[count] == "Suspended":
                    teams.append(0)
                else:
                    teams.append(float(odds_temp[count]))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 4).value = teams[i+1]

    def midasbet(sports):
        #Stupid afl issue with the stupid draw
        if sport!="1":
            print("MidasBet  " + sports_text[int(sport)-1].split(" ")[1])
            #options to access values in each site
            site_text = ["", "", "https://www.midasbet.com.au/Sport/Ice_Hockey/American_NHL/Matches", "https://www.midasbet.com.au/Sport/Basketball/NBA/Matches", "https://www.midasbet.com.au/Sport/Baseball/Major_League_Baseball/Matches", "https://www.midasbet.com.au/Sport/Cricket/Indian_Premier_League/Matches"]
            bets = ["", "", ""]
            driver.get(site_text[int(sport)-1])
            teams = []

            time.sleep(10)

            messy_teams = driver.find_elements(By.XPATH, '//*[@class="teamName"]')


            messy_odds = driver.find_elements(By.XPATH, '//*[@class="betlink oddsColumn"]')

            temp = []
            for y in messy_odds:
                if y.text != '':
                    temp.append(y)
            messy_odds = temp

            if(sport=="4") or (sport=="5") or (sport == "3"):
                temp = []
                for a in messy_odds[::3]:
                    temp.append(a)
                messy_odds = temp
  
            for count, a in enumerate(messy_teams):
                if a.text == "St.Louis Cardinals":
                    teams.append("St. Louis Cardinals")
                else:
                    teams.append(a.text)
                if messy_odds[count].text == "Suspended":
                    teams.append(0)
                else:
                    teams.append(float(messy_odds[count].text))

        
            
            #does the match and then saves values in the right cells
            for i in range(0, len(teams), 2):
                for j in range(0, len(sportsbet_teams)):
                    if teams[i] == sportsbet_teams[j]:
                        sheet.cell(j+m, 5).value = teams[i+1]

    def betright(sports):
        print("Bet Right  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.betright.com.au/sport/basketball/107/united-states-of-america/31/nba/54"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="mantine-Text-root mantine-1q16tia"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="mantine-qo1k2 mantine-Button-label"]')
        print(len(messy_odds))

        temp = []
        temp2 = []
        for y in messy_odds:
            print(y.text)
            if y.text != '':
                temp.append(y)

        print(len(temp))

        for h in temp:
            print(h.text[0])
            if h.text[0].isnumeric():
                print("hello")
                temp2.append(h)
        messy_odds = temp2

        print(len(messy_odds))
        print(len(messy_teams))
        messy_teams.pop()
           
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 6).value = teams[i+1]

    sportsbet(sports)
    #betright(sports)
    bet_deluxe(sports)
    #betr(sports)
    #boombet(sports)
    #midasbet(sports)
    #bet365(sports)
    #bet_nation(sports)
    #unibet(sports)
    #getsetbet(sports)
    #colossalbet(sports)
    #okebet(sports)
    #realbookie(sports)

    workbook.save("Arbitrage.xlsx")

    #Will need to make this a loop for each sheet in the excel
    #Finds which row "Max" is in
    for max_col in range(1, 1048576):
        if sheet.cell(1, max_col).value == "Max":
            break

    #Do the maths ones
    for row in range(m, m + len(sportsbet_teams)):
        #find max
        max_odd = 0
        for col in range(2, max_col):
            if sheet.cell(row, col).value is None:
                sheet.cell(row, col).value = 0
            if sheet.cell(row, col).value > max_odd:
                max_odd = sheet.cell(row, col).value
        sheet.cell(row, max_col).value = max_odd

        #inverse
        sheet.cell(row, max_col+1).value = 1/max_odd

        #finds the sum
        if (sport != "7"):
            if (row-m)%2 != 0 and (row-m) != 0:
                sheet.cell(row, max_col+2).value = sheet.cell(row-1, max_col+1).value + sheet.cell(row, max_col+1).value
                sheet.cell(row-1, max_col+2).value = sheet.cell(row, max_col+2).value
        else:
            if (row-m+1)%3 == 0 and (row-m+1) != 0:
                sheet.cell(row, max_col+2).value = sheet.cell(row-2, max_col+1).value + sheet.cell(row-1, max_col+1).value + sheet.cell(row, max_col+1).value
                sheet.cell(row-1, max_col+2).value = sheet.cell(row, max_col+2).value
                sheet.cell(row-2, max_col+2).value = sheet.cell(row, max_col+2).value 



workbook.save("Arbitrage.xlsx")
driver.quit()

