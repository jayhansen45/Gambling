import requests
import openpyxl as xl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import time
import shutil
import os
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='111.0.5563.64').install(), options = chrome_options)

filedate=date.today()
day = (filedate.strftime('%d/%m/%Y'))
day_site = (filedate.strftime('%A'))
filedate = (filedate.strftime('%A %#d %B %Y'))

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Arbitrage.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\Historical\\Arbitrage.xlsx"
newName = "Arbitrage " + filedate
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\Arbitrage.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\" + newName + ".xlsx")

workbook = xl.load_workbook(filename)

print("Enter the number for the sports you want to check. End with Done")

#Prints the sports options
sports_text = ["1. AFL", "2. NRL", "3. NHL", "4. NBA", "5. MLB"]

for i in sports_text:
    print(i)

#Gets input from the user
sport_input = ""
sports = []

while sport_input != "Done":
    sport_input = input()
    sports.append(sport_input)

sports.pop()

for sport in sports:
    sheet = workbook.worksheets[int(sport)-1]
    
    #Finds the first row that hasn't been used
    m=1

    for m in range(1, 1048576):
        if sheet.cell(m, 1).value is None:
            break

    for start in range(1, 1048576):
        if sheet.cell(start, 3).value is None:
            break
        
    for end in range(1, 1048576):
        if sheet.cell(end, 1).value is None:
            break


    #defines each of the sports as functions
    def sportsbet(sports):
        print("Sportsbet " + sports_text[int(sport)-1].split(" ")[1])

        #options to access values in each site
        site_text = ["https://www.sportsbet.com.au/betting/australian-rules/afl", "https://www.sportsbet.com.au/betting/rugby-league/nrl", "https://www.sportsbet.com.au/betting/ice-hockey-us/nhl-matches", "https://www.sportsbet.com.au/betting/basketball-us/nba", "https://www.sportsbet.com.au/betting/baseball/mlb-matches"]
        bets = ["Head to Head", "Head to Head", "Money Line", "Match Betting", "Money Line"]
        driver.get(site_text[int(sport)-1])
        global sportsbet_teams
        sportsbet_teams = []
        

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[int(sport)-1] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, 'competition-event-participant')]")

        #splits out the team names
        for a in messy_teams[0::3]:
            if (sport == "3") or (sport == "4"):
                sportsbet_teams.append(a.text.split(" @ ")[0])
                sportsbet_teams.append(a.text.split(" @ ")[1])
            elif (sport == "1") or (sport == "2"):
                sportsbet_teams.append(a.text.split(" v ")[0])
                sportsbet_teams.append(a.text.split(" v ")[1])               


        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []

        for a in messy_more_odds:
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))

        for count, teams_list in enumerate(sportsbet_teams):
            sheet.cell(count + m, 1).value = sportsbet_teams[count]
            sheet.cell(count + m, 2).value = more_odds[count]


    def bet_deluxe(sports):
        print("Bet Deluxe  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["https://betdeluxe.com.au/sports/australian-rules/toyota-afl-premiership-1000064", "https://betdeluxe.com.au/sports/rugby-league/telstra-premiership-1000076", "https://betdeluxe.com.au/sports/ice-hockey/national-hockey-league-1000082", "https://betdeluxe.com.au/sports/basketball/nba-1000059"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="errtm9l15 css-115g8u-Text-Text-sportsStyles-styled-sportsStyles__SelectionTitleText-sportsStyles-styled ea6hjv30"]')

        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        stupid_thing = driver.find_elements(By.XPATH, '//*[@class="errtm9l26 css-140o8hh-Text-Text-sportsStyles-styled-sportsStyles__MultiMarketNameText-sportsStyles-styled ea6hjv30"]')

        #Add this back if needed, works for AFl but not for the stupid NRL thingo
        """
        if(sport=="1"):
            temp = []
            for a in messy_odds[::2]:
                temp.append(a)
            messy_odds = temp
        """
        
        if len(sportsbet_teams) != len(messy_teams):
            while len(sportsbet_teams) != len(messy_teams):
                messy_teams.pop()

        for i in messy_odds:
            print(i.text)

        #Can't use popping because then it fuckes up the indexes
        #Work out the indexing maths part and then do a little temp and append like above
        for num, u in enumerate(stupid_thing):
            if u.text == "Handicap (No Draw)":
                print(num)
                messy_odds.pop(2*num)
                messy_odds.pop(2*num+1)
                


        for count, a in enumerate(messy_teams):
            if a.text == "Brisbane Lions":
                teams.append("Brisbane")
            else:
                teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 3).value = teams[i+1]

    #The odds are stored in a stupid secnodary span of the same class?????
    def betr(sports):
        print("BetR  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="OddsButton_priceType__ROL+V SportsMarketCard_priceTypeText__iiWi2"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))
        


        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 3).value = teams[i+1]

    def boombet(sports):
        print("BoomBet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["https://www.boombet.com.au/sport-menu/Sport/Australian%20Rules/AFL", "", "https://www.boombet.com.au/sport-menu/Sport/Ice%20Hockey/US%20NHL%20Regular%20Season-22", "https://www.boombet.com.au/sport-menu/Sport/Basketball/US%20NBA"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="teamName d-block d-md-flex pb-1"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="oddsValue d-block d-md-flex"]')

        #See if this is still required
        """
        if(sport=="1"):
            temp = []
            for a in messy_teams:
                if a.text == "Western Bulldogs":
                    temp.append(a.text)
                else:
                    temp.append(a.text.rsplit(" ", 1)[0])
            messy_teams = temp
        """
                
        for count, a in enumerate(messy_teams):
            if a.text == "Western Bulldogs":
                teams.append(a.text)
            else:
                teams.append(a.text.rsplit(" vs ", 1)[0])
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 4).value = teams[i+1]

    def midasbet(sports):
        print("MidasBet  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.midasbet.com.au/Sport/Basketball/NBA/Matches"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="teamName"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="button betlink button-default"]')

        temp = []
        for y in messy_odds:
            if y.text != '':
                temp.append(y)
        messy_odds = temp

        if(sport=="4"):
            temp = []
            for a in messy_odds[::3]:
                temp.append(a)
            messy_odds = temp
           
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 5).value = teams[i+1]

    def betright(sports):
        print("Bet Right  " + sports_text[int(sport)-1].split(" ")[1])
        #options to access values in each site
        site_text = ["", "", "", "https://www.betright.com.au/sport/basketball/107/united-states-of-america/31/nba/54"]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="mantine-Text-root mantine-1q16tia"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="mantine-qo1k2 mantine-Button-label"]')
        print(len(messy_odds))

        temp = []
        temp2 = []
        for y in messy_odds:
            print(y.text)
            if y.text != '':
                temp.append(y)

        print(len(temp))

        for h in temp:
            print(h.text[0])
            if h.text[0].isnumeric():
                print("hello")
                temp2.append(h)
        messy_odds = temp2

        print(len(messy_odds))
        print(len(messy_teams))
        messy_teams.pop()
           
        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(0, len(sportsbet_teams)):
                if teams[i] == sportsbet_teams[j]:
                    sheet.cell(j+m, 6).value = teams[i+1]

    sportsbet(sports)
    #betright(sports)
    bet_deluxe(sports)
    #betr(sports)
    #boombet(sports)
    #midasbet(sports)

    #Will need to make this a loop for each sheet in the excel
    #Finds which row "Max" is in
    for max_col in range(1, 1048576):
        if sheet.cell(1, max_col).value == "Max":
            break

    #Do the maths ones
    for row in range(m, m + len(sportsbet_teams)):
        #find max
        max_odd = 0
        for col in range(2, max_col):
            if sheet.cell(row, col).value is None:
                sheet.cell(row, col).value = 0
            if sheet.cell(row, col).value > max_odd:
                max_odd = sheet.cell(row, col).value
        sheet.cell(row, max_col).value = max_odd

        #inverse
        sheet.cell(row, max_col+1).value = 1/max_odd

        #finds the sum  
        if (row-m)%2 != 0 and (row-m) != 0:
            sheet.cell(row, max_col+2).value = sheet.cell(row-1, max_col+1).value + sheet.cell(row, max_col+1).value
            sheet.cell(row-1, max_col+2).value = sheet.cell(row, max_col+2).value


workbook.save("Arbitrage.xlsx")
driver.quit()

