import requests
import openpyxl as xl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta, date
import time
import shutil
import os
from selenium.webdriver.common.action_chains import ActionChains

#Bunch of options and shit for the webdriver
chrome_options = webdriver.ChromeOptions()
#chrome_options.binary_location = "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
chrome_options.add_argument('--no-sandbox')
chrome_options.headless = True
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(ChromeDriverManager(version='111.0.5563.64').install(), options = chrome_options)

filedate=date.today()
day = (filedate.strftime('%d/%m/%Y'))
day_site = (filedate.strftime('%A'))
filedate = (filedate.strftime('%A %#d %B %Y'))

#Work Laptop
filename = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Arbitrage.xlsx"
newLocation = "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\Historical\\Arbitrage.xlsx"
newName = "Arbitrage " + filedate
#shutil.copyfile(filename, newLocation)
#os.rename("C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\Arbitrage.xlsx", "C:\\Users\\jhansen3\\OneDrive - KPMG\\Documents\\Python\\Gambling\\Arbitrage\\Historical\\" + newName + ".xlsx")

workbook = xl.load_workbook(filename)
#change below when we do it for every sport
sheet = workbook.worksheets[2]

print("Enter the number for the sports you want to check. End with Done")

#Prints the sports options
sports_text = ["1. AFL", "2. NRL", "3. NHL"]

for i in sports_text:
    print(i)

#Gets input from the user
sport_input = ""
sports = []

while sport_input != "Done":
    sport_input = input()
    sports.append(sport_input)

sports.pop()

#Finds the first row that hasn't been used
m=1

for m in range(1, 1048576):
    if sheet.cell(m, 1).value is None:
        break

for start in range(1, 1048576):
    if sheet.cell(start, 3).value is None:
        break
    
for end in range(1, 1048576):
    if sheet.cell(end, 1).value is None:
        break


#defines each of the sports as functions
def sportsbet(sport):
    #options to access values in each site
    site_text = ["https://www.sportsbet.com.au/betting/australian-rules/afl", "https://www.sportsbet.com.au/betting/rugby-league/nrl", "https://www.sportsbet.com.au/betting/ice-hockey-us/nhl-matches"]
    for sport in sports:
        sheet = workbook.worksheets[int(sport)-1]
        bets = ["Head to Head", "Head to Head", "Money Line"]
        driver.get(site_text[int(sport)-1])
        teams = []

        select = Select(driver.find_element(By.XPATH, '//*[@data-automation-id="market-filter-select"]'))    
        select.select_by_value("Show All Markets")
        time.sleep(10)
        markets = driver.find_elements(By.CLASS_NAME, "market_fjig9r4")
        action = ActionChains(driver)

        for j in range(0, len(markets)):
            if bets[int(sport)-1] == markets[j].text:
                element = markets[j]

        action.click(on_element = element)
        action.perform()
        time.sleep(5)
        messy_teams = driver.find_elements(By.XPATH, "//*[contains(@data-automation-id, 'competition-event-participant')]")

        #splits out the team names
        for a in messy_teams[0::3]:
            teams.append(a.text.split(" @ ")[0])
            teams.append(a.text.split(" @ ")[1])


        messy_more_odds = driver.find_elements(By.XPATH, '//*[@data-automation-id="price-text"]')
        more_odds = []

        for a in messy_more_odds:
            if a.text == "Suspended":
                more_odds.append(0)
            else:
                more_odds.append(float(a.text))

        


        for count, teams_list in enumerate(teams):
            sheet.cell(count + m, 1).value = teams[count]
            sheet.cell(count + m, 2).value = more_odds[count]


def bet_deluxe(sport):
    #options to access values in each site
    site_text = ["", "", "https://betdeluxe.com.au/sports/ice-hockey/national-hockey-league-1000082"]
    for sport in sports:
        sheet = workbook.worksheets[int(sport)-1]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="errtm9l15 css-115g8u-Text-Text-sportsStyles-styled-sportsStyles__SelectionTitleText-sportsStyles-styled ea6hjv30"]')
    
        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(start, end):
                if teams[i] == sheet.cell(j, 1).value:
                    sheet.cell(j, 3).value = teams[i+1]

#The odds are stored in a stupid secnodary span of the same class?????
def betr(sport):
    #options to access values in each site
    site_text = ["", "", "https://betr.com.au/sportsbook#/sport/1006/competition/1000623/1002652"]
    for sport in sports:
        sheet = workbook.worksheets[int(sport)-1]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="OddsButton_priceType__ROL+V SportsMarketCard_priceTypeText__iiWi2"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="errtm9l13 css-16r7ucc-Text-Text-sportsStyles-styled-sportsStyles__OddsText-sportsStyles-styled ea6hjv30"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(start, end):
                if teams[i] == sheet.cell(j, 1).value:
                    sheet.cell(j, 3).value = teams[i+1]

def boombet(sport):
    #options to access values in each site
    site_text = ["", "", "https://www.boombet.com.au/sport-menu/Sport/Ice%20Hockey/US%20NHL%20Regular%20Season-22"]
    for sport in sports:
        sheet = workbook.worksheets[int(sport)-1]
        bets = ["", "", ""]
        driver.get(site_text[int(sport)-1])
        teams = []

        time.sleep(10)

        messy_teams = driver.find_elements(By.XPATH, '//*[@class="teamName d-block d-md-flex pb-1"]')


        messy_odds = driver.find_elements(By.XPATH, '//*[@class="oddsValue d-block d-md-flex"]')

        for count, a in enumerate(messy_teams):
            teams.append(a.text)
            if messy_odds[count].text == "Suspended":
                teams.append(0)
            else:
                teams.append(float(messy_odds[count].text))

        #does the match and then saves values in the right cells
        for i in range(0, len(teams), 2):
            for j in range(start, end):
                if teams[i] == sheet.cell(j, 1).value:
                    sheet.cell(j, 4).value = teams[i+1]

#sportsbet(sports)
bet_deluxe(sports)
#betr(sports)
boombet(sports)

#Will need to make this a loop for each sheet in the excel
#Finds which row "Max" is in
for max_col in range(1, 1048576):
    if sheet.cell(1, max_col).value == "Max":
        break



#Do the maths ones
for row in range(start, end):
    #find max
    max_odd = 0
    for col in range(2, max_col):
        if sheet.cell(row, col).value > max_odd:
            max_odd = sheet.cell(row, col).value
    sheet.cell(row, max_col).value = max_odd

    #inverse
    sheet.cell(row, max_col+1).value = 1/max_odd

    #finds the sum  
    if (row-start)%2 != 0 and (row-start) != 0:
        sheet.cell(row, max_col+2).value = sheet.cell(row-1, max_col+1).value + sheet.cell(row, max_col+1).value
        sheet.cell(row-1, max_col+2).value = sheet.cell(row, max_col+2).value

        

workbook.save("Arbitrage.xlsx")
driver.quit()

